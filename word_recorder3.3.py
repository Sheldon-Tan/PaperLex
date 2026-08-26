import os
import re
import json
import time
import threading
import datetime
import requests
import keyboard
import pyperclip
import tkinter as tk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

EXCEL_FILE = "words.xlsx"

# ==============================================================================
# 🤖 大模型 (LLM) 配置
# ==============================================================================
LLM_CONFIG = {
    "ENABLED": True,
    "API_KEY": "sk-ac925d789ea44ef5bd54e31c74c1c8e5",  # 填入你的 DeepSeek API Key
    "BASE_URL": "https://api.deepseek.com",
    "MODEL": "deepseek-v4-flash",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

NO_PROXY = {"http": None, "https": None}
excel_lock = threading.Lock()


# ==============================================================================
# 📊 Excel 数据存储模块
# ==============================================================================
def init_excel():
    headers = ["单词/内容", "音标", "词性/类型", "完整释义/译文", "记录时间"]
    if not os.path.exists(EXCEL_FILE):
        with excel_lock:
            wb = Workbook()
            ws = wb.active
            ws.title = "生词本"
            ws.append(headers)

            header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for col_num in range(1, 6):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 68
            ws.column_dimensions["E"].width = 18
            wb.save(EXCEL_FILE)


def is_word_already_saved(text: str) -> bool:
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        with excel_lock:
            wb = load_workbook(EXCEL_FILE, read_only=True)
            ws = wb.active
            clean_target = text.strip().lower()
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row and row[0] and str(row[0]).strip().lower() == clean_target:
                    wb.close()
                    return True
            wb.close()
    except Exception:
        pass
    return False


def save_to_excel(info: dict) -> bool:
    try:
        with excel_lock:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            ws.append([
                info["word"],
                info["phonetic"],
                info["pos"],
                info["definition"],
                now
            ])
            last_row = ws.max_row
            ws.cell(row=last_row, column=1).font = Font(name="微软雅黑", size=10, bold=True)
            ws.cell(row=last_row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row=last_row, column=4).font = Font(name="微软雅黑", size=10)
            ws.cell(row=last_row, column=4).alignment = Alignment(wrap_text=True, vertical="center")
            for col_idx in [2, 3, 5]:
                ws.cell(row=last_row, column=col_idx).font = Font(name="微软雅黑", size=10)
                ws.cell(row=last_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
            wb.save(EXCEL_FILE)
            return True
    except Exception:
        return False


# ==============================================================================
# 🔍 查词与语义解析核心
# ==============================================================================
def is_contains_chinese(text: str) -> bool:
    return bool(re.search(r'[\u4e00-\u9fff]', text))


def extract_nested_trans(node):
    results = []
    if isinstance(node, dict):
        if "pos" in node and "tran" in node:
            results.append(f"{node['pos']} {node['tran']}")
        elif "i" in node:
            if isinstance(node["i"], list):
                for item in node["i"]:
                    if isinstance(item, str) and item.strip():
                        results.append(item.strip())
                    elif isinstance(item, dict):
                        results.extend(extract_nested_trans(item))
            elif isinstance(node["i"], str) and node["i"].strip():
                results.append(node["i"].strip())
        elif "tran" in node and isinstance(node["tran"], str):
            results.append(node["tran"].strip())
        else:
            for v in node.values():
                results.extend(extract_nested_trans(v))
    elif isinstance(node, list):
        for item in node:
            results.extend(extract_nested_trans(item))
    return results


def query_youdao(raw_text: str):
    url = "https://dict.youdao.com/jsonapi"
    try:
        response = requests.get(url, params={"client": "mobile", "q": raw_text}, headers=HEADERS, proxies=NO_PROXY,
                                timeout=5)
        if response.status_code == 200:
            return response.json()
    except Exception:
        pass
    return None


def extract_base_form(data: dict, raw_trans: list, clean_text: str):
    if not data:
        return None, None
    for source in ["ec", "simple"]:
        word_obj = data.get(source, {}).get("word", [{}])[0] if data.get(source) else {}
        wfs = word_obj.get("wfs", [])
        for item in wfs:
            wf = item.get("wf", {})
            name = wf.get("name", "")
            val = wf.get("value", "")
            if any(k in name for k in ["原型", "原形", "动词原形", "原型为"]) and val:
                val_clean = val.strip()
                if val_clean.lower() != clean_text.lower():
                    return val_clean, name.strip()

    for source in ["ec", "simple"]:
        word_obj = data.get(source, {}).get("word", [{}])[0] if data.get("source") else {}
        proto = word_obj.get("prototype")
        if proto and isinstance(proto, str) and proto.strip().lower() != clean_text.lower():
            return proto.strip(), "原词"

    pattern = r'[\(（]?([a-zA-Z]+)\s*的(过去式和过去分词|过去分词|过去式|现在分词|复数形式|复数|第三人称单数|比较级|最高级)[\)）]?'
    for line in raw_trans:
        m = re.search(pattern, line)
        if m:
            base_word = m.group(1).strip()
            form_type = m.group(2).strip()
            if base_word.lower() != clean_text.lower():
                return base_word, form_type

    return None, None


def fetch_suggestions(text: str):
    url = "https://dict.youdao.com/suggest"
    try:
        res = requests.get(
            url,
            params={"q": text.strip(), "num": 5, "doctype": "json"},
            headers=HEADERS,
            proxies=NO_PROXY,
            timeout=3
        )
        if res.status_code == 200:
            entries = res.json().get("data", {}).get("entries", [])
            suggestions = []
            for e in entries:
                w = e.get("entry", "").strip()
                exp = e.get("explain", "").strip()
                if w and w.lower() != text.strip().lower():
                    suggestions.append({"word": w, "explain": exp})
            return suggestions
    except Exception:
        pass
    return []


def fetch_free_translation(text: str) -> str:
    data = query_youdao(text)
    if data:
        fanyi = data.get("fanyi", {}).get("tran")
        if fanyi:
            return fanyi.strip()
    return ""


def call_llm_synthesize(phrase: str, word_info_list: list) -> str:
    if not LLM_CONFIG.get("ENABLED"):
        return None
    api_key = LLM_CONFIG.get("API_KEY", "").strip()
    if not api_key or "sk-" not in api_key:
        return None
    base_url = LLM_CONFIG.get("BASE_URL", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"
    parts_desc = [f"{item['word']}({item['definition'].splitlines()[0]})" for item in word_info_list]
    prompt = f"人工智能文献复合词「{phrase}」，分词为：{', '.join(parts_desc)}。请结合 AI/深度学习/计算机语境直接给出最准确的中文释义及词性："
    try:
        payload = {
            "model": LLM_CONFIG.get("MODEL", "deepseek-v4-flash"),
            "messages": [
                {"role": "system", "content": "你是一位专注于人工智能和深度学习学术文献翻译的专家。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.2,
            "max_tokens": 10000
        }
        resp = requests.post(endpoint,
                             headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                             json=payload, timeout=60)
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
    except Exception:
        pass
    return None


def call_llm_direct_query(text: str) -> str:
    """专为人工智能 (AI) 领域论文打造的深度术语解析 (超长生成 & 60秒超时)"""
    if not LLM_CONFIG.get("ENABLED"):
        return "⚠️ LLM 功能未启用，请检查配置。"
    api_key = LLM_CONFIG.get("API_KEY", "").strip()
    if not api_key or "sk-" not in api_key:
        return "⚠️ 未配置有效的 DeepSeek API Key，请在代码顶部填入 sk-xxx。"

    base_url = LLM_CONFIG.get("BASE_URL", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"

    prompt = (
        f"用户正在精读人工智能（AI / 深度学习 / 机器学习 / 大模型 / 计算机视觉 / NLP / 强化学习 等）学术顶会论文，"
        f"遇到了术语、缩写或复合词「{text}」。\n\n"
        f"请结合 AI 学术文献与技术背景，给出专业、详尽、清晰的深度解析：\n"
        f"1. 【核心中文释义】：给出在 AI 领域最准确地道的中文译名、定义及词性。\n"
        f"2. 【全称与技术原理】：若为缩写或学术复合词，给出标准英文全称（如有多种主流 AI 语境展开请分别说明），并详细解析其背后的算法思想、网络结构或数学/技术机制。\n"
        f"3. 【顶会应用语境】：详细说明该术语在 NeurIPS/ICML/CVPR/ACL/ICLR 等顶会论文中通常出现在什么场景（如模型架构、优化算法、损失函数、训练范式等）。\n\n"
        f"请直接输出结构化解析，确保回答详尽完整。"
    )

    try:
        payload = {
            "model": LLM_CONFIG.get("MODEL", "deepseek-v4-flash"),
            "messages": [
                {
                    "role": "system",
                    "content": "你是一位专注于人工智能、机器学习与计算科学领域的资深教授与国际顶会审稿人。请用严谨、专业且详尽的中文解答。"
                },
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.2,
            "max_tokens": 10000
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=60
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
        else:
            return f"⚠️ DeepSeek 请求失败 (状态码: {resp.status_code})"
    except Exception as e:
        return f"⚠️ DeepSeek 网络请求异常: {e}"


def parse_robust_json_candidates(content: str) -> list:
    if not content:
        return []

    clean_json_str = re.sub(r'```(?:json)?', '', content).replace('```', '').strip()
    match = re.search(r'\[.*\]', clean_json_str, re.DOTALL)
    if match:
        try:
            parsed = json.loads(match.group(0))
            if isinstance(parsed, list) and len(parsed) > 0:
                return parsed
        except Exception:
            pass

    candidates = []
    pattern = r'["\']?english["\']?\s*:\s*["\']([^"\']+)["\']\s*,\s*["\']?explain["\']?\s*:\s*["\']([^"\']+)["\']'
    matches = re.findall(pattern, content, re.IGNORECASE)
    for eng, exp in matches:
        candidates.append({"english": eng.strip(), "explain": exp.strip()})

    if candidates:
        return candidates

    lines = content.splitlines()
    for line in lines:
        line_match = re.match(r'^\s*[\d\.\-\*\[\]\(\)]+\s*([a-zA-Z\s\-]+)[\s:：\-\(（]+(.*?)[\)）]?$', line)
        if line_match:
            candidates.append({"english": line_match.group(1).strip(), "explain": line_match.group(2).strip()})

    return candidates


def fetch_english_candidates_from_chinese(chinese_text: str) -> list:
    free_tran = fetch_free_translation(chinese_text)
    api_key = LLM_CONFIG.get("API_KEY", "").strip()

    if not api_key or "sk-" not in api_key or not LLM_CONFIG.get("ENABLED"):
        return [{"english": free_tran, "explain": "免费机器翻译"}] if free_tran else []

    base_url = LLM_CONFIG.get("BASE_URL", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"

    prompt = (
        f"中文「{chinese_text}」，请给出5个常用于人工智能/计算机学术论文或口语日常交流中最贴切的英文表达。\n"
        f"请直接返回严格的 JSON 数组格式，不要包裹多余文字：\n"
        f'[\n'
        f'  {{"english": "expression 1", "explain": "词性/中文释义/学术偏向"}},\n'
        f'  {{"english": "expression 2", "explain": "词性/中文释义/学术偏向"}}\n'
        f']'
    )
    try:
        payload = {
            "model": LLM_CONFIG.get("MODEL", "deepseek-v4-flash"),
            "messages": [
                {"role": "system", "content": "你是一位汉英双语词汇学与 AI 论文写作专家。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.3,
            "max_tokens": 10000
        }
        resp = requests.post(endpoint,
                             headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                             json=payload, timeout=60)
        if resp.status_code == 200:
            content = resp.json()["choices"][0]["message"]["content"].strip()
            results = parse_robust_json_candidates(content)
            if results:
                return results
    except Exception:
        pass

    if free_tran:
        return [{"english": free_tran, "explain": "基础机器翻译 (DeepSeek响应异常时自动兜底)"}]
    return [{"english": chinese_text, "explain": "未能获取到推荐表达"}]


def fetch_single_word_direct(word: str):
    clean_w = word.strip()
    if not clean_w:
        return None
    data = query_youdao(clean_w)
    if not data:
        return None
    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    uk_phone = ec_data.get("ukphone") or simple_data.get("ukphone", "")
    us_phone = ec_data.get("usphone") or simple_data.get("usphone", "")
    phonetic_parts = []
    if uk_phone: phonetic_parts.append(f"英 /{uk_phone}/")
    if us_phone: phonetic_parts.append(f"美 /{us_phone}/")
    if not phonetic_parts:
        phone = ec_data.get("phone") or simple_data.get("phone", "")
        if phone: phonetic_parts.append(f"/{phone}/")
    phonetic_str = " ".join(phonetic_parts) if phonetic_parts else "-"

    raw_trans = []
    for tr in ec_data.get("trs", []):
        for item in extract_nested_trans(tr):
            if item and item not in raw_trans:
                raw_trans.append(item)

    pos_list, final_defs = [], []
    for line in raw_trans:
        m = re.match(r"^([a-zA-Z]+\.)\s*(.+)$", line)
        if m:
            pos = m.group(1)
            if pos not in pos_list: pos_list.append(pos)
            final_defs.append(f"{pos} {m.group(2)}")
        else:
            final_defs.append(line)
    if not final_defs and data.get("fanyi", {}).get("tran"):
        final_defs.append(data["fanyi"]["tran"].strip())

    return {
        "word": clean_w,
        "phonetic": phonetic_str,
        "pos": " / ".join(pos_list) if pos_list else "-",
        "definition": "\n".join(final_defs) if final_defs else "-"
    }


def fetch_entry_info(text: str):
    raw_text = text.strip()
    if not raw_text:
        return None

    clean_text = re.sub(r'[\(（].*?[\)）]', '', raw_text).strip() or raw_text
    candidates = [clean_text]
    if clean_text.lower() not in candidates:
        candidates.append(clean_text.lower())
    if "-" in clean_text:
        candidates.append(clean_text.replace("-", " "))
        candidates.append(clean_text.lower().replace("-", " "))

    data = None
    for cand in candidates:
        res = query_youdao(cand)
        if res and (res.get("ec") or res.get("fanyi") or res.get("web_trans") or res.get("simple")):
            data = res
            break

    split_tokens = [w for w in re.split(r'[-_\s]+', clean_text) if w.strip()]
    if len(split_tokens) >= 2 and (not data or (not data.get("ec") and not data.get("web_trans"))):
        sub_words_info, comb_phonetics = [], []
        for token in split_tokens:
            w_info = fetch_single_word_direct(token)
            if w_info and w_info["definition"] != "-":
                sub_words_info.append(w_info)
                if w_info["phonetic"] != "-":
                    comb_phonetics.append(f"{token}: {w_info['phonetic']}")

        if sub_words_info:
            free_meaning = fetch_free_translation(clean_text.replace("-", " "))
            if not free_meaning or free_meaning.lower() == clean_text.lower():
                free_meaning = call_llm_synthesize(clean_text, sub_words_info) or "无直接组合释义"

            detailed_defs = [f"【AI/推断】: {free_meaning}\n", "【分词详解】:"]
            for sw in sub_words_info:
                sub_def_one_line = sw['definition'].replace('\n', '；')
                detailed_defs.append(f"• {sw['word']} [{sw['pos']}]: {sub_def_one_line}")

            return {
                "word": raw_text,
                "phonetic": " | ".join(comb_phonetics) if comb_phonetics else "-",
                "pos": "[复合词/短语]" if "-" in clean_text else "[短语]",
                "definition": "\n".join(detailed_defs),
                "is_uncertain": False,
                "base_form": None,
                "inflection_type": None
            }

    if not data:
        tran_text = fetch_free_translation(clean_text)
        if tran_text:
            return {
                "word": raw_text,
                "phonetic": "-",
                "pos": "[短语/长句]",
                "definition": tran_text,
                "is_uncertain": False,
                "base_form": None,
                "inflection_type": None
            }
        return None

    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    uk_phone = ec_data.get("ukphone") or simple_data.get("ukphone", "")
    us_phone = ec_data.get("usphone") or simple_data.get("usphone", "")
    phonetic_parts = []
    if uk_phone: phonetic_parts.append(f"英 /{uk_phone}/")
    if us_phone: phonetic_parts.append(f"美 /{us_phone}/")
    if not phonetic_parts:
        phone = ec_data.get("phone") or simple_data.get("phone", "")
        if phone: phonetic_parts.append(f"/{phone}/")
    phonetic_str = "  ".join(phonetic_parts) if phonetic_parts else "-"

    raw_trans = []
    for tr in ec_data.get("trs", []):
        for item in extract_nested_trans(tr):
            if item and item not in raw_trans:
                raw_trans.append(item)

    web_trans_list = []
    for item in data.get("web_trans", {}).get("web-translation", []):
        for t in item.get("trans", []):
            val = t.get("value", "").strip()
            if val and val not in web_trans_list and val not in raw_trans:
                web_trans_list.append(val)

    fanyi_text = data.get("fanyi", {}).get("tran", "").strip()
    words_count = len(clean_text.split())
    final_defs = []
    pos_str = "-"

    if words_count == 1 and "-" not in clean_text:
        pos_list = []
        for line in raw_trans:
            m = re.match(r"^([a-zA-Z]+\.)\s*(.+)$", line)
            if m:
                pos = m.group(1)
                if pos not in pos_list: pos_list.append(pos)
                final_defs.append(f"{pos} {m.group(2)}")
            else:
                final_defs.append(line)
        if not final_defs and web_trans_list:
            final_defs.extend(web_trans_list[:3])
        if not final_defs and fanyi_text:
            final_defs.append(fanyi_text)
        pos_str = " / ".join(pos_list) if pos_list else "-"
    else:
        pos_str = "[复合词]" if "-" in clean_text else "[短语/句子]"
        if raw_trans: final_defs.extend(raw_trans)
        if web_trans_list:
            for w in web_trans_list:
                if w not in final_defs: final_defs.append(w)
        if not final_defs and fanyi_text:
            final_defs.append(fanyi_text)

    is_uncertain = (phonetic_str == "-" and pos_str == "-" and words_count == 1)
    base_form, inflection_type = extract_base_form(data, raw_trans, clean_text)

    return {
        "word": raw_text,
        "phonetic": phonetic_str,
        "pos": pos_str,
        "definition": "\n".join(final_defs) if final_defs else "-",
        "is_uncertain": is_uncertain,
        "base_form": base_form,
        "inflection_type": inflection_type
    }


# ==============================================================================
# 🖥️ 置顶悬浮查词小窗口 (带 DeepSeek 秒级计时与 60s 熔断)
# ==============================================================================
class FloatingLookupApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("📖 生词本(Github:Sheldon-Tan/LinXi) ")
        self.root.geometry("390x320")
        self.root.minsize(320, 250)
        self.root.attributes("-topmost", True)
        self.root.configure(bg="#F8FAFC")

        self.current_info = None

        # ⏱️ 计时器相关状态
        self.timer_running = False
        self.timer_seconds = 0
        self.deepseek_req_id = 0

        self._build_ui()
        self._bind_events()

    def _build_ui(self):
        # 1. 顶部输入搜索栏
        top_frame = tk.Frame(self.root, bg="#F8FAFC", pady=4, padx=8)
        top_frame.pack(fill="x", side="top")

        self.entry_var = tk.StringVar()
        self.search_entry = tk.Entry(
            top_frame,
            textvariable=self.entry_var,
            font=("微软雅黑", 9),
            relief="flat",
            highlightthickness=1,
            highlightbackground="#CBD5E1",
            highlightcolor="#2563EB",
            bg="#FFFFFF"
        )
        self.search_entry.pack(side="left", fill="x", expand=True, ipady=2, padx=(0, 6))

        btn_search = tk.Button(
            top_frame,
            text="查词",
            font=("微软雅黑", 8, "bold"),
            bg="#2563EB",
            fg="white",
            relief="flat",
            activebackground="#1D4ED8",
            activeforeground="white",
            cursor="hand2",
            padx=8,
            pady=1,
            command=self.on_search_click
        )
        btn_search.pack(side="right")

        # 2. 底部操作栏
        bottom_frame = tk.Frame(self.root, bg="#F1F5F9", pady=5, padx=8)
        bottom_frame.pack(fill="x", side="bottom")

        self.lbl_status = tk.Label(bottom_frame, text="[Enter] 存 | [Esc] 隐", font=("微软雅黑", 8), fg="#64748B",
                                   bg="#F1F5F9")
        self.lbl_status.pack(side="left")

        btn_group_frame = tk.Frame(bottom_frame, bg="#F1F5F9")
        btn_group_frame.pack(side="right")

        self.btn_deepseek = tk.Button(
            btn_group_frame,
            text="🤖 DeepSeek",
            font=("微软雅黑", 8, "bold"),
            bg="#6366F1",
            fg="white",
            relief="flat",
            activebackground="#4F46E5",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=2,
            command=self.on_deepseek_click
        )
        self.btn_deepseek.pack(side="left", padx=(0, 4))

        self.btn_save = tk.Button(
            btn_group_frame,
            text="💾 保存 (Enter)",
            font=("微软雅黑", 8, "bold"),
            bg="#10B981",
            fg="white",
            relief="flat",
            activebackground="#059669",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=2,
            command=self.save_current_word
        )
        self.btn_save.pack(side="left")

        # 3. 单词标题 & 音标
        self.header_frame = tk.Frame(self.root, bg="#F8FAFC", padx=8)
        self.header_frame.pack(fill="x", side="top", pady=(1, 2))

        self.lbl_word = tk.Label(self.header_frame, text="就绪", font=("微软雅黑", 11, "bold"), fg="#1E293B",
                                 bg="#F8FAFC")
        self.lbl_word.pack(anchor="w")

        self.lbl_sub = tk.Label(self.header_frame, text="划选按 F8 查词，或在上方打字", font=("微软雅黑", 8),
                                fg="#64748B", bg="#F8FAFC")
        self.lbl_sub.pack(anchor="w")

        # 4. 原词推荐横幅
        self.lemma_frame = tk.Frame(self.root, bg="#EFF6FF", padx=8, pady=2)
        self.lbl_lemma_hint = tk.Label(self.lemma_frame, text="", font=("微软雅黑", 8), fg="#1D4ED8", bg="#EFF6FF")
        self.lbl_lemma_hint.pack(side="left")
        self.btn_lemma_jump = tk.Button(
            self.lemma_frame,
            text="",
            font=("微软雅黑", 8, "bold"),
            bg="#2563EB",
            fg="white",
            relief="flat",
            activebackground="#1D4ED8",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=0
        )
        self.btn_lemma_jump.pack(side="right")

        # 5. 中间内容区 (带滑动条)
        content_frame = tk.Frame(self.root, bg="#F8FAFC", padx=8)
        content_frame.pack(fill="both", expand=True, side="top", pady=2)

        self.cards_frame = tk.Frame(content_frame, bg="#F8FAFC")

        self.txt_frame = tk.Frame(content_frame, bg="#FFFFFF")
        self.txt_frame.pack(fill="both", expand=True)

        self.scrollbar = tk.Scrollbar(self.txt_frame)
        self.scrollbar.pack(side="right", fill="y")

        self.txt_def = tk.Text(
            self.txt_frame,
            font=("微软雅黑", 8),
            bg="#FFFFFF",
            fg="#334155",
            relief="flat",
            highlightthickness=1,
            highlightbackground="#E2E8F0",
            wrap="word",
            padx=6,
            pady=4,
            yscrollcommand=self.scrollbar.set
        )
        self.txt_def.pack(side="left", fill="both", expand=True)
        self.scrollbar.config(command=self.txt_def.yview)

    def _bind_events(self):
        self.search_entry.bind("<Return>", lambda e: self.on_search_click())
        self.root.bind("<Return>", lambda e: self.save_current_word())
        self.root.bind("<Escape>", lambda e: self.hide_window())

    def show_and_focus(self, focus_entry=False):
        self.root.deiconify()
        self.root.attributes("-topmost", True)
        self.root.lift()
        if focus_entry:
            self.search_entry.focus_set()
            self.search_entry.select_range(0, tk.END)

    def hide_window(self):
        self.root.withdraw()

    def set_loading(self, target_text, custom_msg="🔍 正在联网查询中..."):
        self.show_and_focus(focus_entry=False)
        self.entry_var.set(target_text)
        self.lbl_word.config(text=target_text, fg="#1E293B")
        self.lbl_sub.config(text=custom_msg, fg="#2563EB")
        self.lemma_frame.pack_forget()
        self.txt_frame.pack(fill="both", expand=True)
        self.cards_frame.pack_forget()
        self.txt_def.delete("1.0", tk.END)
        self.lbl_status.config(text="正在查询...", fg="#2563EB")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def display_english_result(self, info, suggestions=None):
        self.timer_running = False  # 停止计时
        self.current_info = info
        self.lbl_word.config(text=info["word"], fg="#1E293B")
        sub_text = f"{info['pos']}   {info['phonetic']}"
        self.lbl_sub.config(text=sub_text, fg="#0D9488")

        if info.get("base_form"):
            base_w = info["base_form"]
            inf_type = info.get("inflection_type") or "变形"
            self.lbl_lemma_hint.config(text=f"💡 属「{inf_type}」，原词:")
            self.btn_lemma_jump.config(
                text=f"🔍 查原词 {base_w}",
                command=lambda target=base_w: self.start_lookup(target)
            )
            self.lemma_frame.pack(fill="x", side="top", padx=8, pady=(0, 2), before=self.txt_frame.master)
        else:
            self.lemma_frame.pack_forget()

        self.cards_frame.pack_forget()
        self.txt_frame.pack(fill="both", expand=True)
        self.txt_def.delete("1.0", tk.END)

        display_text = info["definition"]
        if suggestions:
            display_text += "\n\n" + "—" * 20 + "\n💡 相近单词推荐：\n"
            for idx, s in enumerate(suggestions, 1):
                exp = f" ({s['explain']})" if s['explain'] else ""
                display_text += f" • {s['word']}{exp}\n"

        self.txt_def.insert("1.0", display_text)

        if is_word_already_saved(info["word"]):
            self.lbl_status.config(text="💡 已在生词本中", fg="#D97706")
        else:
            self.lbl_status.config(text="👉 按 [Enter] 保存", fg="#64748B")
        self.btn_save.config(state="normal", text="💾 保存 (Enter)", bg="#10B981")

    def display_spelling_suggestions(self, wrong_text, suggestions):
        self.timer_running = False
        self.lemma_frame.pack_forget()
        self.lbl_word.config(text=wrong_text, fg="#DC2626")
        self.lbl_sub.config(text="💡 疑似拼写错误？点击切换：", fg="#D97706")

        self.txt_frame.pack_forget()
        self.cards_frame.pack(fill="both", expand=True)
        for widget in self.cards_frame.winfo_children():
            widget.destroy()

        for idx, item in enumerate(suggestions, 1):
            w = item["word"]
            exp = item["explain"] or "无释义"
            btn = tk.Button(
                self.cards_frame,
                text=f"[{idx}] {w} ({exp})",
                font=("微软雅黑", 8, "bold"),
                bg="#FFFFFF",
                fg="#1E293B",
                relief="flat",
                anchor="w",
                padx=6,
                pady=3,
                cursor="hand2",
                command=lambda target=w: self.start_lookup(target)
            )
            btn.pack(fill="x", pady=2)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#FEF3C7", fg="#D97706"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#FFFFFF", fg="#1E293B"))

        self.lbl_status.config(text="点击推荐词或点右下 DeepSeek", fg="#D97706")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def display_chinese_candidates(self, chinese_text, candidates):
        self.timer_running = False
        self.lemma_frame.pack_forget()
        self.lbl_word.config(text=chinese_text, fg="#1E293B")
        self.lbl_sub.config(text="💡 推荐 5 个 AI 英文表达 (点击选择):", fg="#2563EB")

        self.txt_frame.pack_forget()
        self.cards_frame.pack(fill="both", expand=True)
        for widget in self.cards_frame.winfo_children():
            widget.destroy()

        for idx, item in enumerate(candidates, 1):
            eng = item["english"]
            exp = item["explain"]
            btn = tk.Button(
                self.cards_frame,
                text=f"[{idx}] {eng} ({exp})",
                font=("微软雅黑", 8),
                bg="#FFFFFF",
                fg="#1E293B",
                relief="flat",
                anchor="w",
                padx=6,
                pady=3,
                cursor="hand2",
                command=lambda w=eng: self.start_lookup(w)
            )
            btn.pack(fill="x", pady=1)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#EFF6FF", fg="#2563EB"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#FFFFFF", fg="#1E293B"))

        self.lbl_status.config(text="点击选项获取完整释义", fg="#64748B")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def on_search_click(self):
        text = self.entry_var.get().strip()
        if text:
            self.start_lookup(text)

    def on_deepseek_click(self):
        target = self.entry_var.get().strip()
        if not target and self.current_info:
            target = self.current_info["word"]
        if not target:
            return
        self.start_deepseek_direct_lookup(target)

    def start_lookup(self, text: str):
        self.timer_running = False
        self.set_loading(text)
        threading.Thread(target=self._async_lookup_worker, args=(text,), daemon=True).start()

    def start_deepseek_direct_lookup(self, text: str):
        """启动 DeepSeek 专用通道并激活 60s 动态计时器"""
        self.deepseek_req_id += 1
        req_id = self.deepseek_req_id

        self.timer_running = True
        self.timer_seconds = 0

        self.set_loading(text, custom_msg="🤖 正在请求 DeepSeek 深度解析... (0s / 60s)")
        self._tick_deepseek_timer(req_id)

        threading.Thread(target=self._async_deepseek_worker, args=(text, req_id), daemon=True).start()

    def _tick_deepseek_timer(self, req_id: int):
        """每秒执行一次的动态计时器与 60 秒熔断逻辑"""
        if not self.timer_running or req_id != self.deepseek_req_id:
            return

        self.timer_seconds += 1

        if self.timer_seconds > 60:
            # 🚨 超过 60 秒触发熔断停止
            self.timer_running = False
            self.lbl_sub.config(text="⚠️ 超过 60 秒未响应，AI 已停止", fg="#DC2626")
            self.lbl_status.config(text="AI 已停止", fg="#DC2626")
            self.txt_def.delete("1.0", tk.END)
            self.txt_def.insert(
                "1.0",
                "⚠️ 请求已超过 60 秒，系统已自动停止等待。\n\n"
                "可能原因：\n"
                "1. 当前 DeepSeek 服务器负载较高或正在排队；\n"
                "2. 本地网络连接异常；\n"
                "3. API Key 额度可能不足。\n\n"
                "👉 您可稍后再次点击右下角的 [🤖 DeepSeek] 重试。"
            )
            self.btn_save.config(state="disabled", bg="#94A3B8")
            return

        self.lbl_sub.config(text=f"🤖 正在请求 DeepSeek 深度解析... ({self.timer_seconds}s / 60s)", fg="#2563EB")
        self.root.after(1000, lambda: self._tick_deepseek_timer(req_id))

    def _async_lookup_worker(self, text: str):
        if is_contains_chinese(text):
            candidates = fetch_english_candidates_from_chinese(text)
            self.root.after(0, lambda: self.display_chinese_candidates(text, candidates))
        else:
            info = fetch_entry_info(text)
            suggestions = []
            if not info or info.get("is_uncertain"):
                suggestions = fetch_suggestions(text)

            if not info:
                if suggestions:
                    self.root.after(0, lambda: self.display_spelling_suggestions(text, suggestions))
                else:
                    self.root.after(0, lambda: self._show_not_found(text))
            else:
                self.root.after(0, lambda: self.display_english_result(info, suggestions))

    def _async_deepseek_worker(self, text: str, req_id: int):
        base_info = fetch_entry_info(text)
        phonetic = base_info.get("phonetic", "-") if base_info else "-"
        ai_analysis = call_llm_direct_query(text)

        # 检查是否已超时或被新请求覆盖
        if req_id != self.deepseek_req_id or not self.timer_running:
            return

        self.timer_running = False

        info = {
            "word": text,
            "phonetic": phonetic,
            "pos": "[AI学术]",
            "definition": f"【DeepSeek AI 学术解析】:\n{ai_analysis}",
            "is_uncertain": False,
            "base_form": None,
            "inflection_type": None
        }
        self.root.after(0, lambda: self.display_english_result(info))

    def _show_not_found(self, text: str):
        self.lbl_word.config(text=text, fg="#DC2626")
        self.lbl_sub.config(text="❌ 未查到释义，可点击右下 [DeepSeek] 深度解析", fg="#DC2626")
        self.lbl_status.config(text="词典未收录", fg="#DC2626")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def save_current_word(self):
        if not self.current_info:
            return
        success = save_to_excel(self.current_info)
        if success:
            self.lbl_status.config(text="✨ 已保存至表格！", fg="#059669")
            self.btn_save.config(text="✅ 已保存", state="disabled", bg="#94A3B8")
        else:
            self.lbl_status.config(text="⚠️ 保存失败：请先关闭 Excel 表格！", fg="#DC2626")

    def run(self):
        self.root.mainloop()


# ==============================================================================
# ⌨️ 全局 F8 划词捕获与联动
# ==============================================================================
app_instance = None
is_f8_busy = False


def trigger_f8_hotkey():
    global is_f8_busy, app_instance
    if is_f8_busy or not app_instance:
        return
    is_f8_busy = True
    try:
        pyperclip.copy("")
        time.sleep(0.03)

        keyboard.send("ctrl+c")

        selected_text = ""
        for _ in range(8):
            time.sleep(0.04)
            t = pyperclip.paste().strip()
            if t:
                selected_text = t
                break

        if selected_text:
            app_instance.root.after(0, lambda: app_instance.start_lookup(selected_text))
        else:
            app_instance.root.after(0, lambda: app_instance.show_and_focus(focus_entry=True))
    finally:
        is_f8_busy = False


def start_keyboard_listener():
    keyboard.add_hotkey("f8", trigger_f8_hotkey)
    keyboard.wait()


def main():
    global app_instance
    init_excel()
    print("=" * 60)
    print(" 📖 独立悬浮查词小窗口已启动 (带 DeepSeek 60s 计时与超时熔断)")
    print(" • 【划词查词】：鼠标涂抹选中任何文本 -> 按【F8】自动唤出并查词")
    print(" • 【AI 深度查】：点击右下方【🤖 DeepSeek】开始计时解析 (最长60秒)")
    print(" • 【快捷保存】：按【Enter】保存至表格 | 按【Esc】隐藏窗口")
    print("=" * 60 + "\n")

    t = threading.Thread(target=start_keyboard_listener, daemon=True)
    t.start()

    app_instance = FloatingLookupApp()
    app_instance.run()


if __name__ == "__main__":
    main()