import os
import re
import datetime
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

EXCEL_FILE = "words.xlsx"

# 伪装请求头
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

# 强制直连，防止代理导致网络连接挂起
NO_PROXY = {
    "http": None,
    "https": None
}


def init_excel():
    """初始化 Excel 表格及样式"""
    headers = ["单词/内容", "音标", "词性/类型", "完整释义/译文", "记录时间"]
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "生词本"
        ws.append(headers)

        # 表头格式化：加粗、居中对齐、浅灰色背景
        header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for col_num in range(1, 6):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 预设列宽
        ws.column_dimensions["A"].width = 24  # 单词/内容
        ws.column_dimensions["B"].width = 24  # 音标
        ws.column_dimensions["C"].width = 14  # 词性
        ws.column_dimensions["D"].width = 65  # 释义
        ws.column_dimensions["E"].width = 18  # 记录时间

        wb.save(EXCEL_FILE)


def is_word_already_saved(text: str) -> bool:
    """检查词条是否已经在 Excel 中存在"""
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True)
        ws = wb.active
        clean_target = text.strip().lower()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row and row[0] and str(row[0]).strip().lower() == clean_target:
                wb.close()
                return True
        wb.close()
    except Exception:
        pass
    return False


def extract_nested_trans(node):
    """递归穿透解析词典释义 JSON 节点"""
    results = []
    if isinstance(node, dict):
        if "pos" in node and "tran" in node:
            results.append(f"{node['pos']} {node['tran']}")
        elif "i" in node:
            if isinstance(node["i"], list):
                for item in node["i"]:
                    if isinstance(item, str) and item.strip():
                        results.append(item.strip())
                    elif isinstance(item, dict):
                        results.extend(extract_nested_trans(item))
            elif isinstance(node["i"], str) and node["i"].strip():
                results.append(node["i"].strip())
        elif "tran" in node and isinstance(node["tran"], str):
            results.append(node["tran"].strip())
        else:
            for v in node.values():
                results.extend(extract_nested_trans(v))
    elif isinstance(node, list):
        for item in node:
            results.extend(extract_nested_trans(item))
    return results


def query_youdao(raw_text: str):
    """请求有道底层数据接口"""
    url = "https://dict.youdao.com/jsonapi"
    try:
        response = requests.get(
            url,
            params={"client": "mobile", "q": raw_text},
            headers=HEADERS,
            proxies=NO_PROXY,
            timeout=5
        )
        if response.status_code == 200:
            return response.json()
    except Exception:
        pass
    return None


def get_word_suggestions(raw_text: str, limit: int = 5):
    """获取拼写纠错或相近词推荐列表"""
    suggestions = []
    seen = set()

    # 1. 优先从 jsonapi 的 typos (纠错) 节点读取最精准的修改建议
    data = query_youdao(raw_text)
    if data and "typos" in data:
        typo_items = data["typos"].get("typo", [])
        if isinstance(typo_items, dict):
            typo_items = [typo_items]
        for item in typo_items:
            w = item.get("word", "").strip()
            trans = item.get("trans", "").strip()
            if w and w.lower() != raw_text.lower() and w.lower() not in seen:
                seen.add(w.lower())
                suggestions.append((w, trans))

    # 2. 补充从 suggest 联想接口获取
    if len(suggestions) < limit:
        try:
            res = requests.get(
                "https://dict.youdao.com/suggest",
                params={"q": raw_text, "num": limit + 3, "doctype": "json"},
                headers=HEADERS,
                proxies=NO_PROXY,
                timeout=3
            ).json()
            entries = res.get("data", {}).get("entries", [])
            for item in entries:
                w = item.get("entry", "").strip()
                exp = item.get("explain", "").strip()
                if w and w.lower() != raw_text.lower() and w.lower() not in seen:
                    seen.add(w.lower())
                    suggestions.append((w, exp))
                if len(suggestions) >= limit:
                    break
        except Exception:
            pass

    return suggestions[:limit]


def fetch_single_entry(clean_text: str):
    """单项查询解析主逻辑"""
    candidates = [clean_text]
    if clean_text.lower() not in candidates:
        candidates.append(clean_text.lower())

    data = None
    matched_word = clean_text
    for cand in candidates:
        data = query_youdao(cand)
        if data and (data.get("ec") or data.get("fanyi") or data.get("web_trans") or data.get("simple")):
            matched_word = cand
            break

    if not data:
        return None

    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    # 1. 音标提取
    uk_phone = ec_data.get("ukphone") or simple_data.get("ukphone", "")
    us_phone = ec_data.get("usphone") or simple_data.get("usphone", "")
    phonetic_parts = []
    if uk_phone: phonetic_parts.append(f"英 /{uk_phone}/")
    if us_phone: phonetic_parts.append(f"美 /{us_phone}/")
    if not phonetic_parts:
        phone = ec_data.get("phone") or simple_data.get("phone", "")
        if phone: phonetic_parts.append(f"/{phone}/")
    phonetic_str = "  ".join(phonetic_parts) if phonetic_parts else "-"

    # 2. 词典释义
    raw_trans = []
    for tr in ec_data.get("trs", []):
        for item in extract_nested_trans(tr):
            if item and item not in raw_trans:
                raw_trans.append(item)

    # 3. 网络释义
    web_trans_list = []
    for item in data.get("web_trans", {}).get("web-translation", []):
        for t in item.get("trans", []):
            val = t.get("value", "").strip()
            if val and val not in web_trans_list and val not in raw_trans:
                web_trans_list.append(val)

    # 4. 整句翻译
    fanyi_text = data.get("fanyi", {}).get("tran", "").strip()

    words_count = len(clean_text.split())
    final_defs = []
    pos_str = "-"

    # 场景 A: 单个单词
    if words_count == 1:
        pos_list = []
        for line in raw_trans:
            m = re.match(r"^([a-zA-Z]+\.)\s*(.+)$", line)
            if m:
                pos = m.group(1)
                if pos not in pos_list: pos_list.append(pos)
                final_defs.append(f"{pos} {m.group(2)}")
            else:
                final_defs.append(line)
        if not final_defs and web_trans_list:
            final_defs.extend(web_trans_list[:3])
        if not final_defs and fanyi_text:
            final_defs.append(fanyi_text)
        pos_str = " / ".join(pos_list) if pos_list else "-"

    # 场景 B: 短语 / 搭配
    elif words_count <= 5 and not re.search(r'[.!?。！？]$', clean_text):
        pos_str = "[短语]"
        if raw_trans:
            final_defs.extend(raw_trans)
        if web_trans_list:
            for w in web_trans_list:
                if w not in final_defs: final_defs.append(w)
        if not final_defs and fanyi_text:
            final_defs.append(fanyi_text)

    # 场景 C: 长句子
    else:
        pos_str = "[句子]"
        phonetic_str = "-"
        if fanyi_text:
            final_defs.append(fanyi_text)
        elif raw_trans:
            final_defs.extend(raw_trans)
        elif web_trans_list:
            final_defs.append("；".join(web_trans_list[:3]))

    # 最终兜底
    if not final_defs:
        try:
            s_res = requests.get(
                "https://dict.youdao.com/suggest",
                params={"q": clean_text, "num": 1, "doctype": "json"},
                headers=HEADERS,
                proxies=NO_PROXY,
                timeout=3
            ).json()
            entries = s_res.get("data", {}).get("entries", [])
            if entries and entries[0].get("explain"):
                final_defs.append(entries[0]["explain"])
        except Exception:
            pass

    if not final_defs and phonetic_str == "-":
        return None

    # 提取纠错建议列表
    typos_list = []
    if "typos" in data:
        t_items = data["typos"].get("typo", [])
        if isinstance(t_items, dict): t_items = [t_items]
        for it in t_items:
            tw = it.get("word", "").strip()
            tt = it.get("trans", "").strip()
            if tw and tw.lower() != clean_text.lower():
                typos_list.append((tw, tt))

    return {
        "phonetic": phonetic_str,
        "pos": pos_str,
        "definition": "\n".join(final_defs) if final_defs else "-",
        "suggestions": typos_list
    }


def fetch_entry_info(text: str):
    """入口清洗：处理音标、括号、斜杠多词"""
    raw_text = text.strip()
    if not raw_text:
        return None

    m_phone = re.match(r"^([a-zA-Z\s\-]+)\s*/.*?/$", raw_text)
    if m_phone:
        main_word = m_phone.group(1).strip()
        info = fetch_single_entry(main_word)
        if info:
            info["word"] = raw_text
            return info

    clean_text = re.sub(r'[\(（].*?[\)）]', '', raw_text).strip()
    if not clean_text:
        clean_text = raw_text

    if "/" in clean_text:
        parts = [p.strip() for p in clean_text.split("/") if p.strip()]
        if len(parts) > 1:
            comb_phonetics, comb_pos, comb_defs = [], [], []
            for p in parts:
                res = fetch_single_entry(p)
                if res:
                    if res["phonetic"] != "-": comb_phonetics.append(f"{p}: {res['phonetic']}")
                    if res["pos"] != "-": comb_pos.append(res["pos"])
                    comb_defs.append(f"【{p}】:\n{res['definition']}")
            if comb_defs:
                return {
                    "word": raw_text,
                    "phonetic": " | ".join(comb_phonetics) if comb_phonetics else "-",
                    "pos": " / ".join(dict.fromkeys(comb_pos)) if comb_pos else "[复合项]",
                    "definition": "\n\n".join(comb_defs),
                    "suggestions": []
                }

    info = fetch_single_entry(clean_text)
    if info:
        info["word"] = raw_text
        return info

    return None


def save_to_excel(info: dict):
    """保存并自动应用加粗、换行、居中等样式"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        ws.append([
            info["word"],
            info["phonetic"],
            info["pos"],
            info["definition"],
            now
        ])

        last_row = ws.max_row
        ws.cell(row=last_row, column=1).font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(row=last_row, column=1).alignment = Alignment(wrap_text=True, vertical="center")

        ws.cell(row=last_row, column=4).font = Font(name="微软雅黑", size=10)
        ws.cell(row=last_row, column=4).alignment = Alignment(wrap_text=True, vertical="center")

        for col_idx in [2, 3, 5]:
            ws.cell(row=last_row, column=col_idx).font = Font(name="微软雅黑", size=10)
            ws.cell(row=last_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")

        wb.save(EXCEL_FILE)
        print(f"✨ 已成功将「{info['word']}」保存到 {EXCEL_FILE}\n")
    except PermissionError:
        print(f"⚠️ 保存失败：{EXCEL_FILE} 正在被 Excel/WPS 打开，请先关闭表格后再试！\n")
    except Exception as e:
        print(f"写入文件出错: {e}\n")


def display_and_confirm(info: dict):
    """统一的词条展示与保存确认交互"""
    print("-" * 50)
    print(f"【内容】: {info['word']}")
    print(f"【音标】: {info['phonetic']}")
    print(f"【类型】: {info['pos']}")
    print(f"【释义/译文】:\n{info['definition']}")
    print("-" * 50)

    # 智能判断是否需要展示相近纠错词（若 API 命中 typo，或单词无音标/仅是缩写）
    suggestions = info.get("suggestions", [])
    is_single_word = len(info["word"].split()) == 1
    if not suggestions and is_single_word and (info["phonetic"] == "-" or "abbr." in info["definition"]):
        suggestions = get_word_suggestions(info["word"])

    if suggestions:
        print(f"💡 疑似拼写错误？推荐相近单词：")
        for idx, (s_word, s_exp) in enumerate(suggestions, 1):
            exp_display = f" ({s_exp})" if s_exp else ""
            print(f"   [{idx}] {s_word}{exp_display}")
        print("-" * 50)
        prompt_str = f"👉 [Enter] 保存原词 | 输入序号(1~{len(suggestions)})替换为推荐词 | [N] 取消: "
    else:
        prompt_str = "👉 按 [Enter] 确认保存，输入 [N] 取消保存: "

    try:
        confirm = input(prompt_str).strip()
    except (KeyboardInterrupt, EOFError):
        print("\n已取消。")
        return

    if confirm.lower() in ["n", "no"]:
        print(f"🚫 已跳过「{info['word']}」，未保存到表格。\n")
    elif confirm.isdigit() and suggestions and 1 <= int(confirm) <= len(suggestions):
        selected_word = suggestions[int(confirm) - 1][0]
        print(f"\n🔄 正在重新查询「{selected_word}」...")
        new_info = fetch_entry_info(selected_word)
        if new_info:
            display_and_confirm(new_info)
    else:
        save_to_excel(info)


def main():
    init_excel()
    print("=" * 60)
    print(" 📖 智能生词与短语记录本已就绪 (强化易错词识别)")
    print(" • 查词流程：输入词条 -> (自动检测拼写并推荐正词) -> 确认保存")
    print(" • 操作说明：[Enter] 确认保存 | 输入序号替换为推荐词 | [N] 放弃")
    print(" • 退出方式：输入 q 或按 Ctrl+C 回车退出")
    print("=" * 60 + "\n")

    while True:
        try:
            user_input = input("请输入单词/短语/句子: ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\n程序已安全退出。")
            break

        if not user_input:
            continue
        if user_input.lower() == "q":
            print("程序已退出。")
            break

        if is_word_already_saved(user_input):
            print(f"💡 提示：该词条之前已经记录在 {EXCEL_FILE} 中。")

        print("正在查询中，请稍候...")
        info = fetch_entry_info(user_input)

        if not info:
            suggestions = get_word_suggestions(user_input)
            if suggestions:
                print(f"\n❓ 未找到「{user_input}」的精确释义，您是否想查以下相近单词？")
                for idx, (s_word, s_exp) in enumerate(suggestions, 1):
                    exp_display = f" ({s_exp})" if s_exp else ""
                    print(f"   [{idx}] {s_word}{exp_display}")
                print("-" * 50)
                try:
                    choice = input(f"👉 请输入序号 (1~{len(suggestions)}) 选择正确单词，或输入 [N] 忽略: ").strip()
                except (KeyboardInterrupt, EOFError):
                    print("\n已取消选择。")
                    continue

                if choice.isdigit() and 1 <= int(choice) <= len(suggestions):
                    selected_word = suggestions[int(choice) - 1][0]
                    print(f"\n正在查询「{selected_word}」...")
                    info = fetch_entry_info(selected_word)
                    if info:
                        display_and_confirm(info)
                else:
                    print(f"🚫 已忽略推荐并跳过「{user_input}」。\n")
            else:
                print(f"❌ 未能查到「{user_input}」的相关释义，请检查拼写。\n")
        else:
            display_and_confirm(info)


if __name__ == "__main__":
    main()