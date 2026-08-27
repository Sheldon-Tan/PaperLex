import datetime
import json
import os
import re
import threading
import time
import tkinter as tk
import keyboard
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import pyperclip
import requests

EXCEL_FILE = "words.xlsx"
CONFIG_FILE = "config.json"

DEFAULT_CONFIG = {
    "enabled": True,
    "api_key": "sk-your-api-key",
    "base_url": "https://api.deepseek.com",
    "model": "deepseek-v4-flash"
}


def load_or_init_config() -> dict:
    if not os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(DEFAULT_CONFIG, f, indent=4, ensure_ascii=False)
            print(f"Config initialized: {CONFIG_FILE}")
        except Exception as e:
            print(f"Failed to create {CONFIG_FILE}: {e}")
        return DEFAULT_CONFIG.copy()

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            for key, val in DEFAULT_CONFIG.items():
                if key not in cfg:
                    cfg[key] = val
            return cfg
    except Exception as e:
        print(f"Error reading {CONFIG_FILE}: {e}")
        return DEFAULT_CONFIG.copy()


HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}
NO_PROXY = {"http": None, "https": None}
excel_lock = threading.Lock()


# --- Excel Storage ---

def init_excel():
    headers = ["单词/内容", "音标", "词性/类型", "完整释义/译文", "记录时间"]
    if not os.path.exists(EXCEL_FILE):
        with excel_lock:
            wb = Workbook()
            ws = wb.active
            ws.title = "生词本"
            ws.append(headers)

            header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for col_num in range(1, 6):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 68
            ws.column_dimensions["E"].width = 18
            wb.save(EXCEL_FILE)


def is_word_already_saved(text: str) -> bool:
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        with excel_lock:
            wb = load_workbook(EXCEL_FILE, read_only=True)
            ws = wb.active
            clean_target = text.strip().lower()
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row and row[0] and str(row[0]).strip().lower() == clean_target:
                    wb.close()
                    return True
            wb.close()
    except Exception:
        pass
    return False


def save_to_excel(info: dict) -> bool:
    try:
        with excel_lock:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            ws.append([
                info["word"],
                info["phonetic"],
                info["pos"],
                info["definition"],
                now
            ])
            last_row = ws.max_row
            ws.cell(row=last_row, column=1).font = Font(name="微软雅黑", size=10, bold=True)
            ws.cell(row=last_row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row=last_row, column=4).font = Font(name="微软雅黑", size=10)
            ws.cell(row=last_row, column=4).alignment = Alignment(wrap_text=True, vertical="center")
            for col_idx in [2, 3, 5]:
                ws.cell(row=last_row, column=col_idx).font = Font(name="微软雅黑", size=10)
                ws.cell(row=last_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
            wb.save(EXCEL_FILE)
            return True
    except Exception:
        return False


# --- Dictionary & Parser Helpers ---

def is_contains_chinese(text: str) -> bool:
    return bool(re.search(r'[\u4e00-\u9fff]', text))


def extract_nested_trans(node):
    results = []
    if isinstance(node, dict):
        if "pos" in node and "tran" in node:
            results.append(f"{node['pos']} {node['tran']}")
        elif "i" in node:
            if isinstance(node["i"], list):
                for item in node["i"]:
                    if isinstance(item, str) and item.strip():
                        results.append(item.strip())
                    elif isinstance(item, dict):
                        results.extend(extract_nested_trans(item))
            elif isinstance(node["i"], str) and node["i"].strip():
                results.append(node["i"].strip())
        elif "tran" in node and isinstance(node["tran"], str):
            results.append(node["tran"].strip())
        else:
            for v in node.values():
                results.extend(extract_nested_trans(v))
    elif isinstance(node, list):
        for item in node:
            results.extend(extract_nested_trans(item))
    return results


def query_youdao(raw_text: str):
    url = "https://dict.youdao.com/jsonapi"
    try:
        response = requests.get(
            url,
            params={"client": "mobile", "q": raw_text},
            headers=HEADERS,
            proxies=NO_PROXY,
            timeout=5
        )
        if response.status_code == 200:
            return response.json()
    except Exception:
        pass
    return None


def format_uk_us_verb_variant(val_str: str) -> str:
    if not val_str:
        return ""
    parts = [p.strip() for p in re.split(r'[/,，、或\s]+', val_str) if p.strip()]
    if len(parts) <= 1:
        return val_str

    formatted_parts = []
    for p in parts:
        if re.search(r'll(ed|ing)$', p, re.IGNORECASE):
            formatted_parts.append(f"{p}(英)")
        elif re.search(r'l(ed|ing)$', p, re.IGNORECASE):
            formatted_parts.append(f"{p}(美)")
        elif p.endswith("nt") and any(x.endswith("ned") for x in parts):
            formatted_parts.append(f"{p}(英)")
        elif p.endswith("ned") and any(x.endswith("nt") for x in parts):
            formatted_parts.append(f"{p}(美)")
        elif p.endswith("ised") or p.endswith("ising"):
            formatted_parts.append(f"{p}(英)")
        elif p.endswith("ized") or p.endswith("izing"):
            formatted_parts.append(f"{p}(美)")
        elif "our" in p:
            formatted_parts.append(f"{p}(英)")
        elif "or" in p and any("our" in x for x in parts):
            formatted_parts.append(f"{p}(美)")
        else:
            formatted_parts.append(p)

    seen = set()
    unique_parts = [x for x in formatted_parts if not (x in seen or seen.add(x))]
    return " / ".join(unique_parts)


def extract_verb_forms(data: dict) -> str:
    if not data:
        return ""

    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    wfs_list = ec_data.get("wfs", []) or simple_data.get("wfs", [])
    if not wfs_list:
        return ""

    forms = {}
    for item in wfs_list:
        wf = item.get("wf", {})
        name = wf.get("name", "").strip()
        val = wf.get("value", "").strip()
        if name and val:
            forms[name] = val

    if not any(k in forms for k in ["过去式", "过去分词", "现在分词"]):
        return ""

    verb_items = []
    if "过去式" in forms:
        v = format_uk_us_verb_variant(forms["过去式"])
        verb_items.append(f"过去式: {v}")
    if "过去分词" in forms:
        v = format_uk_us_verb_variant(forms["过去分词"])
        verb_items.append(f"过分: {v}")
    if "现在分词" in forms:
        v = format_uk_us_verb_variant(forms["现在分词"])
        verb_items.append(f"现分: {v}")
    if "第三人称单数" in forms:
        v = format_uk_us_verb_variant(forms["第三人称单数"])
        verb_items.append(f"三单: {v}")

    if len(verb_items) == 4:
        return f"{verb_items[0]}  |  {verb_items[1]}\n{verb_items[2]}  |  {verb_items[3]}"
    elif len(verb_items) > 1:
        return "\n".join(verb_items)
    return verb_items[0] if verb_items else ""


def extract_base_form(data: dict, raw_trans: list, clean_text: str):
    if not data:
        return None, None
    clean_w = clean_text.strip().lower()

    for source in ["ec", "simple"]:
        if not data.get(source):
            continue
        word_obj = data[source].get("word", [{}])[0]
        wfs = word_obj.get("wfs", [])
        for item in wfs:
            wf = item.get("wf", {})
            name = wf.get("name", "")
            val = wf.get("value", "")
            if any(k in name for k in ["原型", "原形", "动词原形", "原型为"]) and val:
                val_clean = val.strip()
                if val_clean.lower() != clean_w:
                    return val_clean, name.strip()

    for source in ["ec", "simple"]:
        if not data.get(source):
            continue
        word_obj = data[source].get("word", [{}])[0]
        proto = word_obj.get("prototype")
        if proto and isinstance(proto, str) and proto.strip().lower() != clean_w:
            return proto.strip(), "原词"

    pattern = r'[\(（]?([a-zA-Z]+)\s*的(过去式和过去分词|过去分词|过去式|现在分词|复数形式|复数|第三人称单数|比较级|最高级)[\)）]?'
    for line in raw_trans:
        m = re.search(pattern, line)
        if m:
            base_word = m.group(1).strip()
            form_type = m.group(2).strip()
            if base_word.lower() != clean_w:
                return base_word, form_type

    candidates = []
    if clean_w.endswith("ia") and len(clean_w) > 3:
        candidates.append((clean_w[:-2] + "ium", "复数 (拉丁原型 -ium)"))
        candidates.append((clean_w[:-2] + "ion", "复数 (希腊原型 -ion)"))
    elif clean_w.endswith("a") and len(clean_w) > 3:
        candidates.append((clean_w[:-1] + "um", "复数 (拉丁原型 -um)"))
        candidates.append((clean_w[:-1] + "on", "复数 (希腊原型 -on)"))
    elif clean_w.endswith("i") and len(clean_w) > 3:
        candidates.append((clean_w[:-1] + "us", "复数 (拉丁原型 -us)"))

    if clean_w.endswith("ing") and len(clean_w) > 4:
        stem = clean_w[:-3]
        candidates.extend([(stem + "e", "现在分词 (-ing)"), (stem, "现在分词 (-ing)")])
        if len(stem) >= 2 and stem[-1] == stem[-2]:
            candidates.append((stem[:-1], "现在分词 (-ing)"))
    elif clean_w.endswith("ed") and len(clean_w) > 3:
        stem = clean_w[:-2]
        candidates.extend([(stem + "e", "过去式/分词 (-ed)"), (stem, "过去式/分词 (-ed)")])
        if clean_w.endswith("ied"):
            candidates.append((clean_w[:-3] + "y", "过去式/分词 (-ied)"))
    elif clean_w.endswith("ies") and len(clean_w) > 4:
        candidates.append((clean_w[:-3] + "y", "复数形式 (-ies)"))
    elif clean_w.endswith("es") and len(clean_w) > 3:
        candidates.append((clean_w[:-2], "复数/三单 (-es)"))
    elif clean_w.endswith("s") and len(clean_w) > 3 and not clean_w.endswith("ss"):
        candidates.append((clean_w[:-1], "复数/三单 (-s)"))

    for cand_word, inflection in candidates:
        if cand_word.lower() != clean_w:
            cand_info = query_youdao(cand_word)
            if cand_info and (cand_info.get("ec") or cand_info.get("simple")):
                return cand_word, inflection

    return None, None


# --- LLM API & Extended Lookups ---

def fetch_antonyms_via_api(word: str) -> str:
    cfg = load_or_init_config()
    if not cfg.get("enabled"):
        return ""
    api_key = cfg.get("api_key", "").strip()
    if not api_key or "sk-" not in api_key:
        return ""

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"

    prompt = f"请给出英文单词「{word}」的 3~6 个最核心英文反义词。直接输出反义词列表并用逗号隔开（例如：word1, word2, word3），不要输出任何中文或多余废话。"
    try:
        payload = {
            "model": cfg.get("model", "deepseek-v4-flash"),
            "messages": [
                {"role": "system", "content": "You are a concise English dictionary. Return only a comma-separated list of antonyms."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.1,
            "max_tokens": 60
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=5
        )
        if resp.status_code == 200:
            content = resp.json()["choices"][0]["message"]["content"].strip()
            return content.replace("\n", " ").replace("`", "").strip()
    except Exception:
        pass
    return ""


def extract_associative_memory(data: dict, clean_w: str = "") -> str:
    if not data and not clean_w:
        return ""

    sections = []

    if data:
        rel_dict = data.get("rel_word", {})
        rel_items = []
        for stem in rel_dict.get("rels", []):
            rel = stem.get("rel", {})
            pos = rel.get("pos", "").strip()
            words = rel.get("words", [])
            pos_prefix = f"[{pos}] " if pos else ""
            w_strs = [f"{w.get('word', '').strip()} ({w.get('tran', '').strip()})" for w in words if w.get('word')]
            if w_strs:
                rel_items.append(f"{pos_prefix}{', '.join(w_strs[:4])}")

        if rel_items:
            sections.append("派生/同源词：\n" + "\n".join(f"  • {item}" for item in rel_items[:4]))

    if data:
        syno_items = []
        syno_dict = data.get("syno", {})
        for item in syno_dict.get("synos", []):
            syn = item.get("syno", {})
            pos = syn.get("pos", "").strip()
            tran = syn.get("tran", "").strip()
            ws_list = [w.get("w", "").strip() for w in syn.get("ws", []) if w.get("w")]
            header = f"[{pos}] {tran}: " if (pos or tran) else ""
            if ws_list:
                syno_items.append(f"{header}{', '.join(ws_list[:5])}")

        if syno_items:
            sections.append("近义词：\n" + "\n".join(f"  • {item}" for item in syno_items[:3]))

    if clean_w:
        anto_str = fetch_antonyms_via_api(clean_w)
        if anto_str:
            sections.append(f"反义词：\n  • {anto_str}")

    return "\n\n".join(sections) if sections else ""


def fetch_suggestions(text: str):
    url = "https://dict.youdao.com/suggest"
    try:
        res = requests.get(
            url,
            params={"q": text.strip(), "num": 5, "doctype": "json"},
            headers=HEADERS,
            proxies=NO_PROXY,
            timeout=3
        )
        if res.status_code == 200:
            entries = res.json().get("data", {}).get("entries", [])
            suggestions = []
            for e in entries:
                w = e.get("entry", "").strip()
                exp = e.get("explain", "").strip()
                if w and w.lower() != text.strip().lower():
                    suggestions.append({"word": w, "explain": exp})
            return suggestions
    except Exception:
        pass
    return []


def fetch_free_translation(text: str) -> str:
    data = query_youdao(text)
    if data:
        fanyi = data.get("fanyi", {}).get("tran")
        if fanyi:
            return fanyi.strip()
    return ""


def call_llm_synthesize(phrase: str, word_info_list: list) -> str:
    cfg = load_or_init_config()
    if not cfg.get("enabled"):
        return None
    api_key = cfg.get("api_key", "").strip()
    if not api_key or "sk-" not in api_key:
        return None

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"
    parts_desc = [f"{item['word']}({item['definition'].splitlines()[0]})" for item in word_info_list]
    prompt = f"人工智能文献复合词「{phrase}」，分词为：{', '.join(parts_desc)}。请结合 AI/深度学习/计算机语境直接给出最准确的中文释义及词性："
    try:
        payload = {
            "model": cfg.get("model", "deepseek-v4-flash"),
            "messages": [
                {"role": "system", "content": "你是一位专注于人工智能和深度学习学术文献翻译的专家。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.2,
            "max_tokens": 10000
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=60
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
    except Exception:
        pass
    return None


def call_llm_direct_query(text: str) -> str:
    cfg = load_or_init_config()
    if not cfg.get("enabled"):
        return "LLM 功能未启用，请在 config.json 中设置 enabled 为 true。"
    api_key = cfg.get("api_key", "").strip()
    if not api_key or "sk-" not in api_key:
        return "未配置有效的 API Key，请在 config.json 中填入 key。"

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"

    prompt = (
        f"用户正在精读人工智能学术顶会论文，遇到术语、单词或缩写「{text}」。\n\n"
        f"请结合 AI 学术文献与语言学背景，给出专业清晰的解析与记忆扩展：\n"
        f"1. 【核心中文释义】：在计算机/AI 领域最准确地道的中文定义、词性及常用场景。\n"
        f"2. 【全称与技术原理】：若为缩写或前沿概念，给出标准英文全称与算法/数学机理。\n"
        f"3. 【顶会应用语境】：说明在 NeurIPS/ICML/CVPR/ICLR 等顶会论文中常出现的场景。\n"
        f"4. 【联想记忆扩展】：给出该词的派生词/同根词 (Derivatives)、常用近义词 (Synonyms) 与 反义词 (Antonyms)，方便学术写作与长效背诵。\n\n"
        f"请直接输出结构化解析，确保排版清晰。"
    )

    try:
        payload = {
            "model": cfg.get("model", "deepseek-v4-flash"),
            "messages": [
                {
                    "role": "system",
                    "content": "你是一位专注于人工智能、机器学习与计算语言学领域的资深学者。请用严谨、专业且详尽的中文解答。"
                },
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.2,
            "max_tokens": 10000
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=60
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
        else:
            return f"请求失败 (状态码: {resp.status_code})"
    except Exception as e:
        return f"网络请求异常: {e}"


def parse_robust_json_candidates(content: str) -> list:
    if not content:
        return []

    clean_json_str = re.sub(r'```(?:json)?', '', content).replace('```', '').strip()
    match = re.search(r'\[.*\]', clean_json_str, re.DOTALL)
    if match:
        try:
            parsed = json.loads(match.group(0))
            if isinstance(parsed, list) and len(parsed) > 0:
                return parsed
        except Exception:
            pass

    candidates = []
    pattern = r'["\']?english["\']?\s*:\s*["\']([^"\']+)["\']\s*,\s*["\']?explain["\']?\s*:\s*["\']([^"\']+)["\']'
    matches = re.findall(pattern, content, re.IGNORECASE)
    for eng, exp in matches:
        candidates.append({"english": eng.strip(), "explain": exp.strip()})

    if candidates:
        return candidates

    lines = content.splitlines()
    for line in lines:
        line_match = re.match(r'^\s*[\d\.\-\*\[\]\(\)]+\s*([a-zA-Z\s\-]+)[\s:：\-\(（]+(.*?)[\)）]?$', line)
        if line_match:
            candidates.append({"english": line_match.group(1).strip(), "explain": line_match.group(2).strip()})

    return candidates


def fetch_english_candidates_from_chinese(chinese_text: str) -> list:
    free_tran = fetch_free_translation(chinese_text)
    cfg = load_or_init_config()
    api_key = cfg.get("api_key", "").strip()

    if not api_key or "sk-" not in api_key or not cfg.get("enabled"):
        return [{"english": free_tran, "explain": "基础机器翻译"}] if free_tran else []

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"

    prompt = (
        f"中文「{chinese_text}」，请给出5个常用于人工智能/计算机学术论文或口语日常交流中最贴切的英文表达。\n"
        f"请直接返回严格的 JSON 数组格式，不要包裹多余文字：\n"
        f'[\n'
        f'  {{"english": "expression 1", "explain": "词性/中文释义/学术偏向"}},\n'
        f'  {{"english": "expression 2", "explain": "词性/中文释义/学术偏向"}}\n'
        f']'
    )
    try:
        payload = {
            "model": cfg.get("model", "deepseek-v4-flash"),
            "messages": [
                {"role": "system", "content": "你是一位汉英双语词汇学与学术写作专家。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.3,
            "max_tokens": 10000
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=60
        )
        if resp.status_code == 200:
            content = resp.json()["choices"][0]["message"]["content"].strip()
            results = parse_robust_json_candidates(content)
            if results:
                return results
    except Exception:
        pass

    if free_tran:
        return [{"english": free_tran, "explain": "基础机器翻译"}]
    return [{"english": chinese_text, "explain": "未能获取到推荐表达"}]


def fetch_single_word_direct(word: str):
    clean_w = word.strip()
    if not clean_w:
        return None
    data = query_youdao(clean_w)
    if not data:
        return None
    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    uk_phone = ec_data.get("ukphone") or simple_data.get("ukphone", "")
    us_phone = ec_data.get("usphone") or simple_data.get("usphone", "")
    phonetic_parts = []
    if uk_phone: phonetic_parts.append(f"英 /{uk_phone}/")
    if us_phone: phonetic_parts.append(f"美 /{us_phone}/")
    if not phonetic_parts:
        phone = ec_data.get("phone") or simple_data.get("phone", "")
        if phone: phonetic_parts.append(f"/{phone}/")
    phonetic_str = " ".join(phonetic_parts) if phonetic_parts else "-"

    raw_trans = []
    for tr in ec_data.get("trs", []):
        for item in extract_nested_trans(tr):
            if item and item not in raw_trans:
                raw_trans.append(item)

    pos_list, final_defs = [], []
    for line in raw_trans:
        m = re.match(r"^([a-zA-Z]+\.)\s*(.+)$", line)
        if m:
            pos = m.group(1)
            if pos not in pos_list: pos_list.append(pos)
            final_defs.append(f"{pos} {m.group(2)}")
        else:
            final_defs.append(line)
    if not final_defs and data.get("fanyi", {}).get("tran"):
        final_defs.append(data["fanyi"]["tran"].strip())

    return {
        "word": clean_w,
        "phonetic": phonetic_str,
        "pos": " / ".join(pos_list) if pos_list else "-",
        "definition": "\n".join(final_defs) if final_defs else "-"
    }


def fetch_entry_info(text: str):
    raw_text = text.strip()
    if not raw_text:
        return None

    clean_text = re.sub(r'[\(（].*?[\)）]', '', raw_text).strip() or raw_text
    candidates = [clean_text]
    if clean_text.lower() not in candidates:
        candidates.append(clean_text.lower())
    if "-" in clean_text:
        candidates.append(clean_text.replace("-", " "))
        candidates.append(clean_text.lower().replace("-", " "))

    data = None
    for cand in candidates:
        res = query_youdao(cand)
        if res and (res.get("ec") or res.get("fanyi") or res.get("web_trans") or res.get("simple")):
            data = res
            break

    split_tokens = [w for w in re.split(r'[-_\s]+', clean_text) if w.strip()]
    if len(split_tokens) >= 2 and (not data or (not data.get("ec") and not data.get("web_trans"))):
        sub_words_info, comb_phonetics = [], []
        for token in split_tokens:
            w_info = fetch_single_word_direct(token)
            if w_info and w_info["definition"] != "-":
                sub_words_info.append(w_info)
                if w_info["phonetic"] != "-":
                    comb_phonetics.append(f"{token}: {w_info['phonetic']}")

        if sub_words_info:
            free_meaning = fetch_free_translation(clean_text.replace("-", " "))
            if not free_meaning or free_meaning.lower() == clean_text.lower():
                free_meaning = call_llm_synthesize(clean_text, sub_words_info) or "无直接组合释义"

            detailed_defs = [f"【推断释义】: {free_meaning}\n", "【分词详解】:"]
            for sw in sub_words_info:
                sub_def_one_line = sw['definition'].replace('\n', '；')
                detailed_defs.append(f"• {sw['word']} [{sw['pos']}]: {sub_def_one_line}")

            return {
                "word": raw_text,
                "phonetic": " | ".join(comb_phonetics) if comb_phonetics else "-",
                "pos": "[复合词/短语]" if "-" in clean_text else "[短语]",
                "definition": "\n".join(detailed_defs),
                "is_uncertain": False,
                "base_form": None,
                "inflection_type": None,
                "verb_forms": "",
                "memory": ""
            }

    if not data:
        tran_text = fetch_free_translation(clean_text)
        if tran_text:
            return {
                "word": raw_text,
                "phonetic": "-",
                "pos": "[短语/长句]",
                "definition": tran_text,
                "is_uncertain": False,
                "base_form": None,
                "inflection_type": None,
                "verb_forms": "",
                "memory": ""
            }
        return None

    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    uk_phone = ec_data.get("ukphone") or simple_data.get("ukphone", "")
    us_phone = ec_data.get("usphone") or simple_data.get("usphone", "")
    phonetic_parts = []
    if uk_phone: phonetic_parts.append(f"英 /{uk_phone}/")
    if us_phone: phonetic_parts.append(f"美 /{us_phone}/")
    if not phonetic_parts:
        phone = ec_data.get("phone") or simple_data.get("phone", "")
        if phone: phonetic_parts.append(f"/{phone}/")
    phonetic_str = "  ".join(phonetic_parts) if phonetic_parts else "-"

    raw_trans = []
    for tr in ec_data.get("trs", []):
        for item in extract_nested_trans(tr):
            if item and item not in raw_trans:
                raw_trans.append(item)

    web_trans_list = []
    for item in data.get("web_trans", {}).get("web-translation", []):
        for t in item.get("trans", []):
            val = t.get("value", "").strip()
            if val and val not in web_trans_list and val not in raw_trans:
                web_trans_list.append(val)

    fanyi_text = data.get("fanyi", {}).get("tran", "").strip()
    words_count = len(clean_text.split())
    final_defs = []
    pos_str = "-"

    if words_count == 1 and "-" not in clean_text:
        pos_list = []
        for line in raw_trans:
            m = re.match(r"^([a-zA-Z]+\.)\s*(.+)$", line)
            if m:
                pos = m.group(1)
                if pos not in pos_list: pos_list.append(pos)
                final_defs.append(f"{pos} {m.group(2)}")
            else:
                final_defs.append(line)
        if not final_defs and web_trans_list:
            final_defs.extend(web_trans_list[:3])
        if not final_defs and fanyi_text:
            final_defs.append(fanyi_text)
        pos_str = " / ".join(pos_list) if pos_list else "-"
    else:
        pos_str = "[复合词]" if "-" in clean_text else "[短语/句子]"
        if raw_trans: final_defs.extend(raw_trans)
        if web_trans_list:
            for w in web_trans_list:
                if w not in final_defs: final_defs.append(w)
        if not final_defs and fanyi_text:
            final_defs.append(fanyi_text)

    is_uncertain = (phonetic_str == "-" and pos_str == "-" and words_count == 1)
    base_form, inflection_type = extract_base_form(data, raw_trans, clean_text)
    verb_forms_str = extract_verb_forms(data)
    memory_str = extract_associative_memory(data, clean_text)

    return {
        "word": raw_text,
        "phonetic": phonetic_str,
        "pos": pos_str,
        "definition": "\n".join(final_defs) if final_defs else "-",
        "is_uncertain": is_uncertain,
        "base_form": base_form,
        "inflection_type": inflection_type,
        "verb_forms": verb_forms_str,
        "memory": memory_str
    }


# --- UI Application ---

class FloatingLookupApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Academic Word Recorder   (Author:Sheldon-Tan)")
        self.root.geometry("410x360")
        self.root.minsize(340, 260)
        self.root.attributes("-topmost", True)
        self.root.configure(bg="#F8FAFC")

        self.current_info = None
        self.timer_running = False
        self.timer_seconds = 0
        self.deepseek_req_id = 0

        self._build_ui()
        self._bind_events()

    def _build_ui(self):
        top_frame = tk.Frame(self.root, bg="#F8FAFC", pady=4, padx=8)
        top_frame.pack(fill="x", side="top")

        self.entry_var = tk.StringVar()
        self.search_entry = tk.Entry(
            top_frame,
            textvariable=self.entry_var,
            font=("微软雅黑", 9),
            relief="flat",
            highlightthickness=1,
            highlightbackground="#CBD5E1",
            highlightcolor="#2563EB",
            bg="#FFFFFF"
        )
        self.search_entry.pack(side="left", fill="x", expand=True, ipady=2, padx=(0, 6))

        btn_search = tk.Button(
            top_frame,
            text="查词",
            font=("微软雅黑", 8, "bold"),
            bg="#2563EB",
            fg="white",
            relief="flat",
            activebackground="#1D4ED8",
            activeforeground="white",
            cursor="hand2",
            padx=8,
            pady=1,
            command=self.on_search_click
        )
        btn_search.pack(side="right")

        bottom_frame = tk.Frame(self.root, bg="#F1F5F9", pady=5, padx=8)
        bottom_frame.pack(fill="x", side="bottom")

        self.lbl_status = tk.Label(bottom_frame, text="[Enter] 存 | [Esc] 隐", font=("微软雅黑", 8), fg="#64748B",
                                   bg="#F1F5F9")
        self.lbl_status.pack(side="left")

        btn_group_frame = tk.Frame(bottom_frame, bg="#F1F5F9")
        btn_group_frame.pack(side="right")

        self.btn_deepseek = tk.Button(
            btn_group_frame,
            text="DeepSeek",
            font=("微软雅黑", 8, "bold"),
            bg="#6366F1",
            fg="white",
            relief="flat",
            activebackground="#4F46E5",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=2,
            command=self.on_deepseek_click
        )
        self.btn_deepseek.pack(side="left", padx=(0, 4))

        self.btn_save = tk.Button(
            btn_group_frame,
            text="保存 (Enter)",
            font=("微软雅黑", 8, "bold"),
            bg="#10B981",
            fg="white",
            relief="flat",
            activebackground="#059669",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=2,
            command=self.save_current_word
        )
        self.btn_save.pack(side="left")

        self.header_frame = tk.Frame(self.root, bg="#F8FAFC", padx=8)
        self.header_frame.pack(fill="x", side="top", pady=(1, 2))

        self.lbl_word = tk.Label(self.header_frame, text="就绪", font=("微软雅黑", 11, "bold"), fg="#1E293B",
                                 bg="#F8FAFC")
        self.lbl_word.pack(anchor="w")

        self.lbl_sub = tk.Label(self.header_frame, text="划选按 F8 查词，或在上方打字", font=("微软雅黑", 8),
                                fg="#64748B", bg="#F8FAFC", justify="left")
        self.lbl_sub.pack(anchor="w")

        self.lemma_frame = tk.Frame(self.root, bg="#EFF6FF", padx=8, pady=2)
        self.lbl_lemma_hint = tk.Label(self.lemma_frame, text="", font=("微软雅黑", 8), fg="#1D4ED8", bg="#EFF6FF")
        self.lbl_lemma_hint.pack(side="left")
        self.btn_lemma_jump = tk.Button(
            self.lemma_frame,
            text="",
            font=("微软雅黑", 8, "bold"),
            bg="#2563EB",
            fg="white",
            relief="flat",
            activebackground="#1D4ED8",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=0
        )
        self.btn_lemma_jump.pack(side="right")

        content_frame = tk.Frame(self.root, bg="#F8FAFC", padx=8)
        content_frame.pack(fill="both", expand=True, side="top", pady=2)

        self.cards_frame = tk.Frame(content_frame, bg="#F8FAFC")
        self.txt_frame = tk.Frame(content_frame, bg="#FFFFFF")
        self.txt_frame.pack(fill="both", expand=True)

        self.scrollbar = tk.Scrollbar(self.txt_frame)
        self.scrollbar.pack(side="right", fill="y")

        self.txt_def = tk.Text(
            self.txt_frame,
            font=("微软雅黑", 8),
            bg="#FFFFFF",
            fg="#334155",
            relief="flat",
            highlightthickness=1,
            highlightbackground="#E2E8F0",
            wrap="word",
            padx=6,
            pady=4,
            yscrollcommand=self.scrollbar.set
        )
        self.txt_def.pack(side="left", fill="both", expand=True)
        self.scrollbar.config(command=self.txt_def.yview)

    def _bind_events(self):
        self.search_entry.bind("<Return>", lambda e: self.on_search_click())
        self.root.bind("<Return>", lambda e: self.save_current_word())
        self.root.bind("<Escape>", lambda e: self.hide_window())

    def show_and_focus(self, focus_entry=False):
        self.root.deiconify()
        self.root.attributes("-topmost", True)
        self.root.lift()
        if focus_entry:
            self.search_entry.focus_set()
            self.search_entry.select_range(0, tk.END)

    def hide_window(self):
        self.root.withdraw()

    def set_loading(self, target_text, custom_msg="正在查询..."):
        self.show_and_focus(focus_entry=False)
        self.entry_var.set(target_text)
        self.lbl_word.config(text=target_text, fg="#1E293B")
        self.lbl_sub.config(text=custom_msg, fg="#2563EB")
        self.lemma_frame.pack_forget()
        self.txt_frame.pack(fill="both", expand=True)
        self.cards_frame.pack_forget()
        self.txt_def.delete("1.0", tk.END)
        self.lbl_status.config(text="正在查询...", fg="#2563EB")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def display_english_result(self, info, suggestions=None):
        self.timer_running = False
        self.current_info = info
        self.lbl_word.config(text=info["word"], fg="#1E293B")

        sub_parts = []
        if info.get("pos") and info["pos"] != "-":
            sub_parts.append(info["pos"])
        if info.get("phonetic") and info["phonetic"] != "-":
            sub_parts.append(info["phonetic"])
        sub_line = "   ".join(sub_parts) if sub_parts else "-"

        if info.get("verb_forms"):
            display_sub = f"{sub_line}\n{info['verb_forms']}"
        else:
            display_sub = sub_line

        self.lbl_sub.config(text=display_sub, fg="#0D9488")

        if info.get("base_form"):
            base_w = info["base_form"]
            inf_type = info.get("inflection_type") or "变形"
            self.lbl_lemma_hint.config(text=f"属「{inf_type}」，原词:")
            self.btn_lemma_jump.config(
                text=f"查原词 {base_w}",
                command=lambda target=base_w: self.start_lookup(target)
            )
            self.lemma_frame.pack(fill="x", side="top", padx=8, pady=(0, 2), before=self.txt_frame.master)
        else:
            self.lemma_frame.pack_forget()

        self.cards_frame.pack_forget()
        self.txt_frame.pack(fill="both", expand=True)
        self.txt_def.delete("1.0", tk.END)

        display_text = info["definition"]

        if info.get("memory"):
            display_text += "\n\n" + "—" * 12 + " 联想记忆 " + "—" * 12 + "\n" + info["memory"]

        if suggestions:
            display_text += "\n\n" + "—" * 12 + " 相近推荐 " + "—" * 12 + "\n"
            for idx, s in enumerate(suggestions, 1):
                exp = f" ({s['explain']})" if s['explain'] else ""
                display_text += f" • {s['word']}{exp}\n"

        self.txt_def.insert("1.0", display_text)

        if is_word_already_saved(info["word"]):
            self.lbl_status.config(text="已在生词本中", fg="#D97706")
        else:
            self.lbl_status.config(text="按 [Enter] 保存", fg="#64748B")
        self.btn_save.config(state="normal", text="保存 (Enter)", bg="#10B981")

    def display_spelling_suggestions(self, wrong_text, suggestions):
        self.timer_running = False
        self.lemma_frame.pack_forget()
        self.lbl_word.config(text=wrong_text, fg="#DC2626")
        self.lbl_sub.config(text="疑似拼写错误，点击切换：", fg="#D97706")

        self.txt_frame.pack_forget()
        self.cards_frame.pack(fill="both", expand=True)
        for widget in self.cards_frame.winfo_children():
            widget.destroy()

        for idx, item in enumerate(suggestions, 1):
            w = item["word"]
            exp = item["explain"] or "无释义"
            btn = tk.Button(
                self.cards_frame,
                text=f"[{idx}] {w} ({exp})",
                font=("微软雅黑", 8, "bold"),
                bg="#FFFFFF",
                fg="#1E293B",
                relief="flat",
                anchor="w",
                padx=6,
                pady=3,
                cursor="hand2",
                command=lambda target=w: self.start_lookup(target)
            )
            btn.pack(fill="x", pady=2)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#FEF3C7", fg="#D97706"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#FFFFFF", fg="#1E293B"))

        self.lbl_status.config(text="点击推荐词或点右下 DeepSeek", fg="#D97706")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def display_chinese_candidates(self, chinese_text, candidates):
        self.timer_running = False
        self.lemma_frame.pack_forget()
        self.lbl_word.config(text=chinese_text, fg="#1E293B")
        self.lbl_sub.config(text="推荐英文表达 (点击选择):", fg="#2563EB")

        self.txt_frame.pack_forget()
        self.cards_frame.pack(fill="both", expand=True)
        for widget in self.cards_frame.winfo_children():
            widget.destroy()

        for idx, item in enumerate(candidates, 1):
            eng = item["english"]
            exp = item["explain"]
            btn = tk.Button(
                self.cards_frame,
                text=f"[{idx}] {eng} ({exp})",
                font=("微软雅黑", 8),
                bg="#FFFFFF",
                fg="#1E293B",
                relief="flat",
                anchor="w",
                padx=6,
                pady=3,
                cursor="hand2",
                command=lambda w=eng: self.start_lookup(w)
            )
            btn.pack(fill="x", pady=1)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#EFF6FF", fg="#2563EB"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#FFFFFF", fg="#1E293B"))

        self.lbl_status.config(text="点击选项获取完整释义", fg="#64748B")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def on_search_click(self):
        text = self.entry_var.get().strip()
        if text:
            self.start_lookup(text)

    def on_deepseek_click(self):
        target = self.entry_var.get().strip()
        if not target and self.current_info:
            target = self.current_info["word"]
        if not target:
            return
        self.start_deepseek_direct_lookup(target)

    def start_lookup(self, text: str):
        self.timer_running = False
        self.set_loading(text)
        threading.Thread(target=self._async_lookup_worker, args=(text,), daemon=True).start()

    def start_deepseek_direct_lookup(self, text: str):
        self.deepseek_req_id += 1
        req_id = self.deepseek_req_id

        self.timer_running = True
        self.timer_seconds = 0

        self.set_loading(text, custom_msg="正在请求 DeepSeek 解析... (0s / 60s)")
        self._tick_deepseek_timer(req_id)

        threading.Thread(target=self._async_deepseek_worker, args=(text, req_id), daemon=True).start()

    def _tick_deepseek_timer(self, req_id: int):
        if not self.timer_running or req_id != self.deepseek_req_id:
            return

        self.timer_seconds += 1

        if self.timer_seconds > 60:
            self.timer_running = False
            self.lbl_sub.config(text="请求超时 (60s)，已停止", fg="#DC2626")
            self.lbl_status.config(text="超时停止", fg="#DC2626")
            self.txt_def.delete("1.0", tk.END)
            self.txt_def.insert(
                "1.0",
                "请求超过 60 秒未响应，已自动中止。\n"
                "可能原因：\n"
                "1. 网络连接不稳定；\n"
                "2. 接口负载过高；\n"
                "3. API Key 或配置无效。"
            )
            self.btn_save.config(state="disabled", bg="#94A3B8")
            return

        self.lbl_sub.config(text=f"正在请求 DeepSeek 解析... ({self.timer_seconds}s / 60s)", fg="#2563EB")
        self.root.after(1000, lambda: self._tick_deepseek_timer(req_id))

    def _async_lookup_worker(self, text: str):
        if is_contains_chinese(text):
            candidates = fetch_english_candidates_from_chinese(text)
            self.root.after(0, lambda: self.display_chinese_candidates(text, candidates))
        else:
            info = fetch_entry_info(text)
            suggestions = []
            if not info or info.get("is_uncertain"):
                suggestions = fetch_suggestions(text)

            if not info:
                if suggestions:
                    self.root.after(0, lambda: self.display_spelling_suggestions(text, suggestions))
                else:
                    self.root.after(0, lambda: self._show_not_found(text))
            else:
                self.root.after(0, lambda: self.display_english_result(info, suggestions))

    def _async_deepseek_worker(self, text: str, req_id: int):
        base_info = fetch_entry_info(text)
        phonetic = base_info.get("phonetic", "-") if base_info else "-"
        verb_forms = base_info.get("verb_forms", "") if base_info else ""
        ai_analysis = call_llm_direct_query(text)

        if req_id != self.deepseek_req_id or not self.timer_running:
            return

        self.timer_running = False

        info = {
            "word": text,
            "phonetic": phonetic,
            "pos": "[学术解析]",
            "definition": f"【学术解析】:\n{ai_analysis}",
            "is_uncertain": False,
            "base_form": None,
            "inflection_type": None,
            "verb_forms": verb_forms,
            "memory": ""
        }
        self.root.after(0, lambda: self.display_english_result(info))

    def _show_not_found(self, text: str):
        self.lbl_word.config(text=text, fg="#DC2626")
        self.lbl_sub.config(text="词典未收录，可尝试 DeepSeek 深度解析", fg="#DC2626")
        self.lbl_status.config(text="未收录", fg="#DC2626")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def save_current_word(self):
        if not self.current_info:
            return
        success = save_to_excel(self.current_info)
        if success:
            self.lbl_status.config(text="已保存至表格", fg="#059669")
            self.btn_save.config(text="已保存", state="disabled", bg="#94A3B8")
        else:
            self.lbl_status.config(text="保存失败：请检查 Excel 是否被占用", fg="#DC2626")

    def run(self):
        self.root.mainloop()


# --- Global Listener ---

app_instance = None
is_f8_busy = False


def trigger_f8_hotkey():
    global is_f8_busy, app_instance
    if is_f8_busy or not app_instance:
        return
    is_f8_busy = True
    try:
        pyperclip.copy("")
        time.sleep(0.03)

        keyboard.send("ctrl+c")

        selected_text = ""
        for _ in range(8):
            time.sleep(0.04)
            t = pyperclip.paste().strip()
            if t:
                selected_text = t
                break

        if selected_text:
            app_instance.root.after(0, lambda: app_instance.start_lookup(selected_text))
        else:
            app_instance.root.after(0, lambda: app_instance.show_and_focus(focus_entry=True))
    finally:
        is_f8_busy = False


def start_keyboard_listener():
    keyboard.add_hotkey("f8", trigger_f8_hotkey)
    keyboard.wait()


def main():
    global app_instance
    init_excel()
    load_or_init_config()

    print(f"Loaded config: {CONFIG_FILE}")
    print("Service running. Hotkey [F8] enabled.")

    t = threading.Thread(target=start_keyboard_listener, daemon=True)
    t.start()

    app_instance = FloatingLookupApp()
    app_instance.run()


if __name__ == "__main__":
    main()