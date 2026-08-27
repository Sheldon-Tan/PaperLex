#优化了deepseek提示词,并增加了Latex渲染功能
# PaperLex - 学术论文精读桌面词典 (支持高清 LaTeX 公式光栅化与 Markdown 原生表格渲染)
import datetime
import io
import json
import os
import re
import threading
import time
import tkinter as tk
from tkinter import messagebox, simpledialog
import keyboard
import matplotlib

matplotlib.use("Agg")  # 强制无头绘图模式
import matplotlib.pyplot as plt
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import pyperclip
import requests

EXCEL_FILE = "words.xlsx"
CONFIG_FILE = "config.json"

DEFAULT_CONFIG = {
    "enabled": True,
    "api_key": "sk-your-api-key",
    "base_url": "https://api.deepseek.com",
    "model": "deepseek-chat"
}


# --- Configuration Management ---

def load_or_init_config() -> dict:
    if not os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(DEFAULT_CONFIG, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Failed to create {CONFIG_FILE}: {e}")
        return DEFAULT_CONFIG.copy()

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            for key, val in DEFAULT_CONFIG.items():
                if key not in cfg:
                    cfg[key] = val
            return cfg
    except Exception as e:
        print(f"Error reading {CONFIG_FILE}: {e}")
        return DEFAULT_CONFIG.copy()


def save_config(cfg: dict) -> bool:
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Failed to save {CONFIG_FILE}: {e}")
        return False


def open_settings_dialog(parent_window):
    cfg = load_or_init_config()
    current_key = cfg.get("api_key", "").strip()
    if current_key == "sk-your-api-key" or "sk-你的" in current_key:
        current_key = ""

    new_key = simpledialog.askstring(
        "API Key 配置",
        "请输入 DeepSeek API Key (sk-xxxx)：\n\n"
        "• 仅保存在本地 config.json\n"
        "• 留空不影响基础查词、动词时态与派生词拓展",
        initialvalue=current_key,
        show="*",
        parent=parent_window
    )

    if new_key is not None:
        cfg["api_key"] = new_key.strip()
        cfg["enabled"] = bool(new_key.strip())
        save_config(cfg)
        if new_key.strip():
            messagebox.showinfo("成功", "API Key 已成功保存", parent=parent_window)


HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}
NO_PROXY = {"http": None, "https": None}
excel_lock = threading.Lock()


# --- Ultra-HD LaTeX & Markdown Rich Text Engine ---

def sanitize_latex_for_mathtext(latex_expr: str) -> str:
    """自动清理并修复 Matplotlib MathText 不支持的宏命令"""
    clean_expr = latex_expr.strip()
    clean_expr = re.sub(r'^(?:\\\[|\$\$|\\\(|\$)|(?:\\\]|\$\$|\\\)|\$)$', '', clean_expr).strip()

    # 移除 MathText 会崩溃的括号大小修饰符 (\big, \Big, \bigg, \Bigg 等)
    clean_expr = re.sub(r'\\(?:big|Big|bigg|Bigg)[lrm]?', '', clean_expr)

    # 宏命令替换为 MathText 标准支持形式
    clean_expr = re.sub(r'\\operatorname\{([^}]+)\}', r'\\mathrm{\1}', clean_expr)
    clean_expr = re.sub(r'\\text\{([^}]+)\}', r'\\mathrm{\1}', clean_expr)
    clean_expr = re.sub(r'\\bm\{([^}]+)\}', r'\\mathbf{\1}', clean_expr)
    clean_expr = re.sub(r'\\boldsymbol\{([^}]+)\}', r'\\mathbf{\1}', clean_expr)

    # 处理虚拟边界点
    clean_expr = clean_expr.replace(r'\left.', '').replace(r'\right.', '')

    return clean_expr.strip()


def render_latex_image(latex_expr: str, is_block: bool = True):
    """采用 300 DPI 超采样 + Lanczos 重采样滤波，生成矢量级高清渲染图"""
    clean_expr = sanitize_latex_for_mathtext(latex_expr)

    fig = plt.figure(figsize=(0.01, 0.01))
    try:
        fontsize = 12 if is_block else 9.5
        fig.text(0, 0, f"${clean_expr}$", fontsize=fontsize, color="#0F172A")

        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=300, bbox_inches="tight", pad_inches=0.02, transparent=True)
        buf.seek(0)

        img = Image.open(buf)
        scale_ratio = 0.38
        target_w = max(1, int(img.width * scale_ratio))
        target_h = max(1, int(img.height * scale_ratio))
        img_sharp = img.resize((target_w, target_h), Image.Resampling.LANCZOS)

        return ImageTk.PhotoImage(img_sharp)
    finally:
        plt.close(fig)


def setup_rich_text_tags(text_widget: tk.Text):
    """注册 Markdown 各级富文本标签样式"""
    text_widget.tag_configure("h1", font=("微软雅黑", 10, "bold"), foreground="#0F172A", spacing1=6, spacing3=2)
    text_widget.tag_configure("h2", font=("微软雅黑", 9, "bold"), foreground="#1E293B", spacing1=5, spacing3=2)
    text_widget.tag_configure("h3", font=("微软雅黑", 8, "bold"), foreground="#2563EB", spacing1=4, spacing3=1)
    text_widget.tag_configure("bold", font=("微软雅黑", 8, "bold"), foreground="#0F172A")
    text_widget.tag_configure("normal", font=("微软雅黑", 8), foreground="#334155")
    text_widget.tag_configure("bullet", font=("微软雅黑", 8, "bold"), foreground="#2563EB")
    text_widget.tag_configure("hr", font=("微软雅黑", 6), foreground="#CBD5E1")
    text_widget.tag_configure("quote", font=("微软雅黑", 8, "italic"), foreground="#64748B")


def parse_inline_tokens(text_widget: tk.Text, line_str: str, base_tag: str = "normal"):
    """解析行内元素（**加粗** 与 \\(行内公式\\)）"""
    pattern = r'(\*\*.*?\*\*|\\\(.*?\\\)|\$.*?\$)'
    parts = re.split(pattern, line_str)

    for part in parts:
        if not part:
            continue
        if part.startswith("**") and part.endswith("**") and len(part) >= 4:
            bold_text = part[2:-2]
            text_widget.insert(tk.END, bold_text, "bold")
        elif (part.startswith(r"\(") and part.endswith(r"\)")) or (
                part.startswith("$") and part.endswith("$") and len(part) > 1):
            try:
                tk_img = render_latex_image(part, is_block=False)
                text_widget._images_cache.append(tk_img)
                text_widget.image_create(tk.END, image=tk_img)
            except Exception:
                text_widget.insert(tk.END, part, base_tag)
        else:
            text_widget.insert(tk.END, part, base_tag)


def render_markdown_table(text_widget: tk.Text, table_lines: list):
    """将 Markdown 表格文本动态渲染为原生嵌入式 GUI 表格组件"""
    if len(table_lines) < 2:
        return

    parsed_rows = []
    for line in table_lines:
        row = [cell.strip() for cell in line.strip().strip('|').split('|')]
        # 跳过分割线行 (如 |---|---|)
        if all(re.match(r'^:?-+:?$', cell) for cell in row if cell):
            continue
        if any(row):
            parsed_rows.append(row)

    if not parsed_rows:
        return

    col_count = max(len(r) for r in parsed_rows)

    # 创建表格容器
    tbl_frame = tk.Frame(text_widget, bg="#CBD5E1", bd=1, relief="solid")
    text_widget._table_frames.append(tbl_frame)

    for r_idx, row in enumerate(parsed_rows):
        is_header = (r_idx == 0)
        bg_color = "#F1F5F9" if is_header else ("#FFFFFF" if r_idx % 2 == 1 else "#F8FAFC")
        fg_color = "#0F172A" if is_header else "#334155"
        font_style = ("微软雅黑", 8, "bold") if is_header else ("微软雅黑", 8)

        for c_idx in range(col_count):
            cell_text = row[c_idx] if c_idx < len(row) else ""
            cell_lbl = tk.Label(
                tbl_frame,
                text=cell_text,
                font=font_style,
                bg=bg_color,
                fg=fg_color,
                padx=6,
                pady=3,
                anchor="w",
                justify="left"
            )
            cell_lbl.grid(row=r_idx, column=c_idx, sticky="nsew", padx=0.5, pady=0.5)

    for c in range(col_count):
        tbl_frame.grid_columnconfigure(c, weight=1)

    text_widget.insert(tk.END, "\n")
    text_widget.window_create(tk.END, window=tbl_frame)
    text_widget.insert(tk.END, "\n\n")


def render_markdown_with_latex(text_widget: tk.Text, raw_text: str):
    """Markdown 语法、原生表格与 LaTeX 公式的完整图文混排解析器"""
    if not hasattr(text_widget, "_images_cache"):
        text_widget._images_cache = []
    if not hasattr(text_widget, "_table_frames"):
        text_widget._table_frames = []

    text_widget._images_cache.clear()
    text_widget._table_frames.clear()
    setup_rich_text_tags(text_widget)

    # 1. 优先拆分出独立的跨行公式块 \\[ ... \\] 或 $$ ... $$
    block_math_pattern = r'(\\\[[\s\S]*?\\\]|\$\$[\s\S]*?\$\$)'
    sections = re.split(block_math_pattern, raw_text)

    for section in sections:
        if not section:
            continue

        is_block_math = (section.startswith(r"\[") and section.endswith(r"\]")) or (
                section.startswith("$$") and section.endswith("$$"))

        if is_block_math:
            try:
                tk_img = render_latex_image(section, is_block=True)
                text_widget._images_cache.append(tk_img)
                text_widget.insert(tk.END, "\n")
                text_widget.image_create(tk.END, image=tk_img)
                text_widget.insert(tk.END, "\n\n")
            except Exception:
                text_widget.insert(tk.END, f"\n{section}\n", "normal")
        else:
            # 2. 处理常规文本与 Markdown 表格块
            lines = section.split("\n")
            table_buffer = []
            in_table = False

            for idx, line in enumerate(lines):
                trimmed = line.strip()

                # 判断是否为表格行
                if trimmed.startswith("|") and trimmed.endswith("|") and len(trimmed) > 1:
                    in_table = True
                    table_buffer.append(trimmed)
                    continue
                else:
                    if in_table and table_buffer:
                        render_markdown_table(text_widget, table_buffer)
                        table_buffer = []
                        in_table = False

                if not trimmed:
                    if idx < len(lines) - 1:
                        text_widget.insert(tk.END, "\n")
                    continue

                if trimmed.startswith("---") or trimmed.startswith("___"):
                    text_widget.insert(tk.END, "─" * 46 + "\n", "hr")
                elif trimmed.startswith("# "):
                    parse_inline_tokens(text_widget, trimmed[2:], base_tag="h1")
                    text_widget.insert(tk.END, "\n")
                elif trimmed.startswith("## "):
                    parse_inline_tokens(text_widget, trimmed[3:], base_tag="h2")
                    text_widget.insert(tk.END, "\n")
                elif trimmed.startswith("### "):
                    parse_inline_tokens(text_widget, trimmed[4:], base_tag="h3")
                    text_widget.insert(tk.END, "\n")
                elif trimmed.startswith("- ") or trimmed.startswith("* "):
                    text_widget.insert(tk.END, " • ", "bullet")
                    parse_inline_tokens(text_widget, line.lstrip()[2:], base_tag="normal")
                    text_widget.insert(tk.END, "\n")
                elif trimmed.startswith("> "):
                    text_widget.insert(tk.END, " ▌ ", "bullet")
                    parse_inline_tokens(text_widget, line.lstrip()[2:], base_tag="quote")
                    text_widget.insert(tk.END, "\n")
                else:
                    parse_inline_tokens(text_widget, line, base_tag="normal")
                    text_widget.insert(tk.END, "\n")

            if in_table and table_buffer:
                render_markdown_table(text_widget, table_buffer)


# --- Excel Storage ---

def init_excel():
    headers = ["单词/内容", "音标", "词性/类型", "完整释义/译文", "记录时间"]
    if not os.path.exists(EXCEL_FILE):
        with excel_lock:
            wb = Workbook()
            ws = wb.active
            ws.title = "生词本"
            ws.append(headers)

            header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for col_num in range(1, 6):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 68
            ws.column_dimensions["E"].width = 18
            wb.save(EXCEL_FILE)


def is_word_already_saved(text: str) -> bool:
    if not os.path.exists(EXCEL_FILE):
        return False
    try:
        with excel_lock:
            wb = load_workbook(EXCEL_FILE, read_only=True)
            ws = wb.active
            clean_target = text.strip().lower()
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row and row[0] and str(row[0]).strip().lower() == clean_target:
                    wb.close()
                    return True
            wb.close()
    except Exception:
        pass
    return False


def save_to_excel(info: dict) -> bool:
    try:
        with excel_lock:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            ws.append([
                info["word"],
                info["phonetic"],
                info["pos"],
                info["definition"],
                now
            ])
            last_row = ws.max_row
            ws.cell(row=last_row, column=1).font = Font(name="微软雅黑", size=10, bold=True)
            ws.cell(row=last_row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row=last_row, column=4).font = Font(name="微软雅黑", size=10)
            ws.cell(row=last_row, column=4).alignment = Alignment(wrap_text=True, vertical="center")
            for col_idx in [2, 3, 5]:
                ws.cell(row=last_row, column=col_idx).font = Font(name="微软雅黑", size=10)
                ws.cell(row=last_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
            wb.save(EXCEL_FILE)
            return True
    except Exception:
        return False


# --- Dictionary & Parser Core ---

def is_contains_chinese(text: str) -> bool:
    return bool(re.search(r'[\u4e00-\u9fff]', text))


def extract_nested_trans(node):
    results = []
    if isinstance(node, dict):
        if "pos" in node and "tran" in node:
            results.append(f"{node['pos']} {node['tran']}")
        elif "i" in node:
            if isinstance(node["i"], list):
                for item in node["i"]:
                    if isinstance(item, str) and item.strip():
                        results.append(item.strip())
                    elif isinstance(item, dict):
                        results.extend(extract_nested_trans(item))
            elif isinstance(node["i"], str) and node["i"].strip():
                results.append(node["i"].strip())
        elif "tran" in node and isinstance(node["tran"], str):
            results.append(node["tran"].strip())
        else:
            for v in node.values():
                results.extend(extract_nested_trans(v))
    elif isinstance(node, list):
        for item in node:
            results.extend(extract_nested_trans(item))
    return results


def query_youdao(raw_text: str):
    url = "https://dict.youdao.com/jsonapi"
    try:
        response = requests.get(
            url,
            params={"client": "mobile", "q": raw_text},
            headers=HEADERS,
            proxies=NO_PROXY,
            timeout=5
        )
        if response.status_code == 200:
            return response.json()
    except Exception:
        pass
    return None


def format_uk_us_verb_variant(val_str: str) -> str:
    if not val_str:
        return ""
    parts = [p.strip() for p in re.split(r'[/,，、或\s]+', val_str) if p.strip()]
    if len(parts) <= 1:
        return val_str

    formatted_parts = []
    for p in parts:
        if re.search(r'll(ed|ing)$', p, re.IGNORECASE):
            formatted_parts.append(f"{p}(英)")
        elif re.search(r'l(ed|ing)$', p, re.IGNORECASE):
            formatted_parts.append(f"{p}(美)")
        elif p.endswith("nt") and any(x.endswith("ned") for x in parts):
            formatted_parts.append(f"{p}(英)")
        elif p.endswith("ned") and any(x.endswith("nt") for x in parts):
            formatted_parts.append(f"{p}(美)")
        elif p.endswith("ised") or p.endswith("ising"):
            formatted_parts.append(f"{p}(英)")
        elif p.endswith("ized") or p.endswith("izing"):
            formatted_parts.append(f"{p}(美)")
        elif "our" in p:
            formatted_parts.append(f"{p}(英)")
        elif "or" in p and any("our" in x for x in parts):
            formatted_parts.append(f"{p}(美)")
        else:
            formatted_parts.append(p)

    seen = set()
    unique_parts = [x for x in formatted_parts if not (x in seen or seen.add(x))]
    return " / ".join(unique_parts)


def extract_verb_forms(data: dict) -> str:
    if not data:
        return ""

    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    wfs_list = ec_data.get("wfs", []) or simple_data.get("wfs", [])
    if not wfs_list:
        return ""

    forms = {}
    for item in wfs_list:
        wf = item.get("wf", {})
        name = wf.get("name", "").strip()
        val = wf.get("value", "").strip()
        if name and val:
            forms[name] = val

    if not any(k in forms for k in ["过去式", "过去分词", "现在分词"]):
        return ""

    verb_items = []
    if "过去式" in forms:
        v = format_uk_us_verb_variant(forms["过去式"])
        verb_items.append(f"过去式: {v}")
    if "过去分词" in forms:
        v = format_uk_us_verb_variant(forms["过去分词"])
        verb_items.append(f"过分: {v}")
    if "现在分词" in forms:
        v = format_uk_us_verb_variant(forms["现在分词"])
        verb_items.append(f"现分: {v}")
    if "第三人称单数" in forms:
        v = format_uk_us_verb_variant(forms["第三人称单数"])
        verb_items.append(f"三单: {v}")

    if len(verb_items) == 4:
        return f"{verb_items[0]}  |  {verb_items[1]}\n{verb_items[2]}  |  {verb_items[3]}"
    elif len(verb_items) > 1:
        return "\n".join(verb_items)
    return verb_items[0] if verb_items else ""


def extract_base_form(data: dict, raw_trans: list, clean_text: str):
    if not data:
        return None, None
    clean_w = clean_text.strip().lower()

    for source in ["ec", "simple"]:
        if not data.get(source):
            continue
        word_obj = data[source].get("word", [{}])[0]
        wfs = word_obj.get("wfs", [])
        for item in wfs:
            wf = item.get("wf", {})
            name = wf.get("name", "")
            val = wf.get("value", "")
            if any(k in name for k in ["原型", "原形", "动词原形", "原型为"]) and val:
                val_clean = val.strip()
                if val_clean.lower() != clean_w:
                    return val_clean, name.strip()

    for source in ["ec", "simple"]:
        if not data.get(source):
            continue
        word_obj = data[source].get("word", [{}])[0]
        proto = word_obj.get("prototype")
        if proto and isinstance(proto, str) and proto.strip().lower() != clean_w:
            return proto.strip(), "原词"

    pattern = r'[\(（]?([a-zA-Z]+)\s*的(过去式和过去分词|过去分词|过去式|现在分词|复数形式|复数|第三人称单数|比较级|最高级)[\)）]?'
    for line in raw_trans:
        m = re.search(pattern, line)
        if m:
            base_word = m.group(1).strip()
            form_type = m.group(2).strip()
            if base_word.lower() != clean_w:
                return base_word, form_type

    candidates = []
    if clean_w.endswith("ia") and len(clean_w) > 3:
        candidates.append((clean_w[:-2] + "ium", "复数 (拉丁原型 -ium)"))
        candidates.append((clean_w[:-2] + "ion", "复数 (希腊原型 -ion)"))
    elif clean_w.endswith("a") and len(clean_w) > 3:
        candidates.append((clean_w[:-1] + "um", "复数 (拉丁原型 -um)"))
        candidates.append((clean_w[:-1] + "on", "复数 (希腊原型 -on)"))
    elif clean_w.endswith("i") and len(clean_w) > 3:
        candidates.append((clean_w[:-1] + "us", "复数 (拉丁原型 -us)"))

    if clean_w.endswith("ing") and len(clean_w) > 4:
        stem = clean_w[:-3]
        candidates.extend([(stem + "e", "现在分词 (-ing)"), (stem, "现在分词 (-ing)")])
        if len(stem) >= 2 and stem[-1] == stem[-2]:
            candidates.append((stem[:-1], "现在分词 (-ing)"))
    elif clean_w.endswith("ed") and len(clean_w) > 3:
        stem = clean_w[:-2]
        candidates.extend([(stem + "e", "过去式/分词 (-ed)"), (stem, "过去式/分词 (-ed)")])
        if clean_w.endswith("ied"):
            candidates.append((clean_w[:-3] + "y", "过去式/分词 (-ied)"))
    elif clean_w.endswith("ies") and len(clean_w) > 4:
        candidates.append((clean_w[:-3] + "y", "复数形式 (-ies)"))
    elif clean_w.endswith("es") and len(clean_w) > 3:
        candidates.append((clean_w[:-2], "复数/三单 (-es)"))
    elif clean_w.endswith("s") and len(clean_w) > 3 and not clean_w.endswith("ss"):
        candidates.append((clean_w[:-1], "复数/三单 (-s)"))

    for cand_word, inflection in candidates:
        if cand_word.lower() != clean_w:
            cand_info = query_youdao(cand_word)
            if cand_info and (cand_info.get("ec") or cand_info.get("simple")):
                return cand_word, inflection

    return None, None


# --- LLM API & Extended Lookups ---

def probe_llm_disambiguation(text: str) -> list:
    """第一阶段（极速嗅探，仅耗~50 Token）：判断是否为多义学术缩写/术语"""
    clean_t = text.strip()
    if len(clean_t.split()) > 2 or "(" in clean_t or len(clean_t) > 15:
        return []

    cfg = load_or_init_config()
    if not cfg.get("enabled"):
        return []
    api_key = cfg.get("api_key", "").strip()
    if not api_key or "sk-" not in api_key:
        return []

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"
    model_name = cfg.get("model", "deepseek-chat")

    prompt = (
        f"判断学术词汇/缩写「{clean_t}」在计算机视觉、语音、NLP、多模态、机器学习或交叉AI学术领域是否存在2个以上不同的主流学术全称/研究课题。\n"
        f"若是多义词，请返回2~4个最核心全称及领域的严格 JSON 数组，严禁包含任何其他多余文本：\n"
        f'[\n'
        f'  {{"full_name": "Active Speaker Detection", "domain": "音视频多模态 / 说话人检测"}},\n'
        f'  {{"full_name": "Autism Spectrum Disorder", "domain": "医学AI / 情感计算 / 脑机"}}\n'
        f']\n'
        f"如果该词无多义性或只是普通英文单词，请直接返回空数组: []"
    )
    try:
        payload = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": "You are a concise academic disambiguation detector. Output JSON only."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.1,
            "max_tokens": 200
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=10
        )
        if resp.status_code == 200:
            content = resp.json()["choices"][0]["message"]["content"].strip()
            clean_json_str = re.sub(r'```(?:json)?', '', content).replace('```', '').strip()
            match = re.search(r'\[.*\]', clean_json_str, re.DOTALL)
            if match:
                parsed = json.loads(match.group(0))
                if isinstance(parsed, list) and len(parsed) > 1:
                    return parsed
    except Exception:
        pass
    return []


def fetch_antonyms_via_api(word: str) -> str:
    cfg = load_or_init_config()
    if not cfg.get("enabled"):
        return ""
    api_key = cfg.get("api_key", "").strip()
    if not api_key or "sk-" not in api_key:
        return ""

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"

    prompt = f"请给出英文单词「{word}」的 3~6 个最核心英文反义词。直接输出反义词列表并用逗号隔开（例如：word1, word2, word3），不要输出任何中文或多余废话。"
    try:
        payload = {
            "model": cfg.get("model", "deepseek-chat"),
            "messages": [
                {"role": "system",
                 "content": "You are a concise English dictionary. Return only a comma-separated list of antonyms."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.1,
            "max_tokens": 120
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=5
        )
        if resp.status_code == 200:
            content = resp.json()["choices"][0]["message"]["content"].strip()
            return content.replace("\n", " ").replace("`", "").strip()
    except Exception:
        pass
    return ""


def extract_associative_memory(data: dict, clean_w: str = "") -> str:
    if not data and not clean_w:
        return ""

    sections = []

    if data:
        rel_dict = data.get("rel_word", {})
        rel_items = []
        for stem in rel_dict.get("rels", []):
            rel = stem.get("rel", {})
            pos = rel.get("pos", "").strip()
            words = rel.get("words", [])
            pos_prefix = f"[{pos}] " if pos else ""
            w_strs = [f"{w.get('word', '').strip()} ({w.get('tran', '').strip()})" for w in words if w.get('word')]
            if w_strs:
                rel_items.append(f"{pos_prefix}{', '.join(w_strs[:4])}")

        if rel_items:
            sections.append("### 派生/同源词\n" + "\n".join(f"- {item}" for item in rel_items[:4]))

    if data:
        syno_items = []
        syno_dict = data.get("syno", {})
        for item in syno_dict.get("synos", []):
            syn = item.get("syno", {})
            pos = syn.get("pos", "").strip()
            tran = syn.get("tran", "").strip()
            ws_list = [w.get("w", "").strip() for w in syn.get("ws", []) if w.get("w")]
            header = f"[{pos}] {tran}: " if (pos or tran) else ""
            if ws_list:
                syno_items.append(f"{header}{', '.join(ws_list[:5])}")

        if syno_items:
            sections.append("### 近义词\n" + "\n".join(f"- {item}" for item in syno_items[:3]))

    if clean_w:
        anto_str = fetch_antonyms_via_api(clean_w)
        if anto_str:
            sections.append(f"### 反义词\n- {anto_str}")

    return "\n\n".join(sections) if sections else ""


def fetch_suggestions(text: str):
    url = "https://dict.youdao.com/suggest"
    try:
        res = requests.get(
            url,
            params={"q": text.strip(), "num": 5, "doctype": "json"},
            headers=HEADERS,
            proxies=NO_PROXY,
            timeout=3
        )
        if res.status_code == 200:
            entries = res.json().get("data", {}).get("entries", [])
            suggestions = []
            for e in entries:
                w = e.get("entry", "").strip()
                exp = e.get("explain", "").strip()
                if w and w.lower() != text.strip().lower():
                    suggestions.append({"word": w, "explain": exp})
            return suggestions
    except Exception:
        pass
    return []


def fetch_free_translation(text: str) -> str:
    data = query_youdao(text)
    if data:
        fanyi = data.get("fanyi", {}).get("tran")
        if fanyi:
            return fanyi.strip()
    return ""


def call_llm_synthesize(phrase: str, word_info_list: list) -> str:
    cfg = load_or_init_config()
    if not cfg.get("enabled"):
        return None
    api_key = cfg.get("api_key", "").strip()
    if not api_key or "sk-" not in api_key:
        return None

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"
    parts_desc = [f"{item['word']}({item['definition'].splitlines()[0]})" for item in word_info_list]
    prompt = f"人工智能文献复合词「{phrase}」，分词为：{', '.join(parts_desc)}。请结合 AI/深度学习/计算机语境直接给出最准确的中文释义及词性："
    try:
        payload = {
            "model": cfg.get("model", "deepseek-chat"),
            "messages": [
                {"role": "system", "content": "你是一位专注于人工智能和深度学习学术文献翻译的专家。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.2,
            "max_tokens": 1000
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=60
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
    except Exception:
        pass
    return None


def call_llm_direct_query(text: str, context_hint: str = "") -> str:
    """第二阶段：携带精准消歧上下文的顶会学术深度推演引擎"""
    cfg = load_or_init_config()
    if not cfg.get("enabled"):
        return "LLM 功能未启用，请在配置中开启。"
    api_key = cfg.get("api_key", "").strip()
    if not api_key or "sk-" not in api_key:
        return "未配置有效的 API Key，请点击右下角设置配置。"

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"

    model_name = cfg.get("model", "deepseek-chat")
    if model_name in ["deepseek-v4-flash", "default"]:
        model_name = "deepseek-chat"

    target_desc = f"「{context_hint}」" if context_hint else f"「{text}」"

    prompt = (
        f"你在协助用户精读计算机视觉(CV)、自然语言处理(NLP)、语音与多模态AI领域的顶级学术论文（CVPR/ICCV/ECCV/NeurIPS/ICLR/Interspeech等）。\n"
        f"用户选中的学术目标为：{target_desc}。\n\n"
        f"【公式排版硬性约束】：所有跨行独立数学公式必须单独成行并严格使用 \\[ 与 \\] 包裹；行内简短数学变量使用 \\( 与 \\) 包裹。\n\n"
        f"请自适应判断目标类型并输出学术干货：\n\n"
        f"【类型 A：技术任务 / 前沿概念 / 专业缩写（如 Active Speaker Detection, CLIP, LoRA, DETR 等）】\n"
        f"### 1. 📌 标准中文术语名\n学术界公认中文规范定名与英文全称。\n\n"
        f"### 2. 🔬 核心技术机理\n输入输出模态、损失函数/数学原理与核心架构。\n\n"
        f"### 3. 📊 顶会语境与基准\n包含经典数据集与代表性 SOTA 基线表格对比（请使用标准 Markdown 表格输出）。\n\n"
        f"### 4. 💡 关联上下游技术\n2~3 个常一同出现的关联学术概念（辨析区别）。\n\n"
        f"【类型 B：常规学术英语词汇 / 动词 / 形容词】\n"
        f"### 1. 📖 学术释义与词性\n科研论文中最常见地道的含义与语境偏向。\n\n"
        f"### 2. 📝 顶会论文典型例句\n1 句高水平顶会论文中的地道用法及中文对照。\n\n"
        f"### 3. 🧠 联想近义词与派生同根词\n\n"
        f"请直接输出符合上述标准格式的 Markdown 内容。"
    )

    try:
        payload = {
            "model": model_name,
            "messages": [
                {
                    "role": "system",
                    "content": "你是一位计算机视觉、语音信号处理与多模态人工智能领域的顶会审稿人与资深学者。解答严谨、专业、直击技术要点。"
                },
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.1,
            "max_tokens": 2500
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=60
        )
        if resp.status_code == 200:
            return resp.json()["choices"][0]["message"]["content"].strip()
        else:
            return f"请求失败 (状态码: {resp.status_code})"
    except Exception as e:
        return f"网络请求异常: {e}"


def parse_robust_json_candidates(content: str) -> list:
    if not content:
        return []

    clean_json_str = re.sub(r'```(?:json)?', '', content).replace('```', '').strip()
    match = re.search(r'\[.*\]', clean_json_str, re.DOTALL)
    if match:
        try:
            parsed = json.loads(match.group(0))
            if isinstance(parsed, list) and len(parsed) > 0:
                return parsed
        except Exception:
            pass

    candidates = []
    pattern = r'["\']?english["\']?\s*:\s*["\']([^"\']+)["\']\s*,\s*["\']?explain["\']?\s*:\s*["\']([^"\']+)["\']'
    matches = re.findall(pattern, content, re.IGNORECASE)
    for eng, exp in matches:
        candidates.append({"english": eng.strip(), "explain": exp.strip()})

    if candidates:
        return candidates

    lines = content.splitlines()
    for line in lines:
        line_match = re.match(r'^\s*[\d\.\-\*\[\]\(\)]+\s*([a-zA-Z\s\-]+)[\s:：\-\(（]+(.*?)[\)）]?$', line)
        if line_match:
            candidates.append({"english": line_match.group(1).strip(), "explain": line_match.group(2).strip()})

    return candidates


def fetch_english_candidates_from_chinese(chinese_text: str) -> list:
    free_tran = fetch_free_translation(chinese_text)
    cfg = load_or_init_config()
    api_key = cfg.get("api_key", "").strip()

    if not api_key or "sk-" not in api_key or not cfg.get("enabled"):
        return [{"english": free_tran, "explain": "基础机器翻译"}] if free_tran else []

    base_url = cfg.get("base_url", "https://api.deepseek.com").rstrip("/")
    endpoint = f"{base_url}/chat/completions" if base_url.endswith("/v1") else f"{base_url}/v1/chat/completions"

    prompt = (
        f"中文「{chinese_text}」，请给出5个常用于人工智能/计算机学术论文或口语日常交流中最贴切的英文表达。\n"
        f"请直接返回严格的 JSON 数组格式，不要包裹多余文字：\n"
        f'[\n'
        f'  {{"english": "expression 1", "explain": "词性/中文释义/学术偏向"}},\n'
        f'  {{"english": "expression 2", "explain": "词性/中文释义/学术偏向"}}\n'
        f']'
    )
    try:
        payload = {
            "model": cfg.get("model", "deepseek-chat"),
            "messages": [
                {"role": "system", "content": "你是一位汉英双语词汇学与学术写作专家。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.3,
            "max_tokens": 1000
        }
        resp = requests.post(
            endpoint,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=60
        )
        if resp.status_code == 200:
            content = resp.json()["choices"][0]["message"]["content"].strip()
            results = parse_robust_json_candidates(content)
            if results:
                return results
    except Exception:
        pass

    if free_tran:
        return [{"english": free_tran, "explain": "基础机器翻译"}]
    return [{"english": chinese_text, "explain": "未能获取到推荐表达"}]


def fetch_single_word_direct(word: str):
    clean_w = word.strip()
    if not clean_w:
        return None
    data = query_youdao(clean_w)
    if not data:
        return None
    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    uk_phone = ec_data.get("ukphone") or simple_data.get("ukphone", "")
    us_phone = ec_data.get("usphone") or simple_data.get("usphone", "")
    phonetic_parts = []
    if uk_phone: phonetic_parts.append(f"英 /{uk_phone}/")
    if us_phone: phonetic_parts.append(f"美 /{us_phone}/")
    if not phonetic_parts:
        phone = ec_data.get("phone") or simple_data.get("phone", "")
        if phone: phonetic_parts.append(f"/{phone}/")
    phonetic_str = " ".join(phonetic_parts) if phonetic_parts else "-"

    raw_trans = []
    for tr in ec_data.get("trs", []):
        for item in extract_nested_trans(tr):
            if item and item not in raw_trans:
                raw_trans.append(item)

    pos_list, final_defs = [], []
    for line in raw_trans:
        m = re.match(r"^([a-zA-Z]+\.)\s*(.+)$", line)
        if m:
            pos = m.group(1)
            if pos not in pos_list: pos_list.append(pos)
            final_defs.append(f"**{pos}** {m.group(2)}")
        else:
            final_defs.append(line)
    if not final_defs and data.get("fanyi", {}).get("tran"):
        final_defs.append(data["fanyi"]["tran"].strip())

    return {
        "word": clean_w,
        "phonetic": phonetic_str,
        "pos": " / ".join(pos_list) if pos_list else "-",
        "definition": "\n".join(final_defs) if final_defs else "-"
    }


def fetch_entry_info(text: str):
    raw_text = text.strip()
    if not raw_text:
        return None

    clean_text = re.sub(r'[\(（].*?[\)）]', '', raw_text).strip() or raw_text
    candidates = [clean_text]
    if clean_text.lower() not in candidates:
        candidates.append(clean_text.lower())
    if "-" in clean_text:
        candidates.append(clean_text.replace("-", " "))
        candidates.append(clean_text.lower().replace("-", " "))

    data = None
    for cand in candidates:
        res = query_youdao(cand)
        if res and (res.get("ec") or res.get("fanyi") or res.get("web_trans") or res.get("simple")):
            data = res
            break

    split_tokens = [w for w in re.split(r'[-_\s]+', clean_text) if w.strip()]
    if len(split_tokens) >= 2 and (not data or (not data.get("ec") and not data.get("web_trans"))):
        sub_words_info, comb_phonetics = [], []
        for token in split_tokens:
            w_info = fetch_single_word_direct(token)
            if w_info and w_info["definition"] != "-":
                sub_words_info.append(w_info)
                if w_info["phonetic"] != "-":
                    comb_phonetics.append(f"{token}: {w_info['phonetic']}")

        if sub_words_info:
            free_meaning = fetch_free_translation(clean_text.replace("-", " "))
            if not free_meaning or free_meaning.lower() == clean_text.lower():
                free_meaning = call_llm_synthesize(clean_text, sub_words_info) or "无直接组合释义"

            detailed_defs = [f"**【推断释义】**: {free_meaning}\n", "### 分词详解"]
            for sw in sub_words_info:
                sub_def_one_line = sw['definition'].replace('\n', '；')
                detailed_defs.append(f"- **{sw['word']}** `[{sw['pos']}]`: {sub_def_one_line}")

            return {
                "word": raw_text,
                "phonetic": " | ".join(comb_phonetics) if comb_phonetics else "-",
                "pos": "[复合词/短语]" if "-" in clean_text else "[短语]",
                "definition": "\n".join(detailed_defs),
                "is_uncertain": False,
                "base_form": None,
                "inflection_type": None,
                "verb_forms": "",
                "memory": ""
            }

    if not data:
        tran_text = fetch_free_translation(clean_text)
        if tran_text:
            return {
                "word": raw_text,
                "phonetic": "-",
                "pos": "[短语/长句]",
                "definition": tran_text,
                "is_uncertain": False,
                "base_form": None,
                "inflection_type": None,
                "verb_forms": "",
                "memory": ""
            }
        return None

    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    uk_phone = ec_data.get("ukphone") or simple_data.get("ukphone", "")
    us_phone = ec_data.get("usphone") or simple_data.get("usphone", "")
    phonetic_parts = []
    if uk_phone: phonetic_parts.append(f"英 /{uk_phone}/")
    if us_phone: phonetic_parts.append(f"美 /{us_phone}/")
    if not phonetic_parts:
        phone = ec_data.get("phone") or simple_data.get("phone", "")
        if phone: phonetic_parts.append(f"/{phone}/")
    phonetic_str = "  ".join(phonetic_parts) if phonetic_parts else "-"

    raw_trans = []
    for tr in ec_data.get("trs", []):
        for item in extract_nested_trans(tr):
            if item and item not in raw_trans:
                raw_trans.append(item)

    web_trans_list = []
    for item in data.get("web_trans", {}).get("web-translation", []):
        for t in item.get("trans", []):
            val = t.get("value", "").strip()
            if val and val not in web_trans_list and val not in raw_trans:
                web_trans_list.append(val)

    fanyi_text = data.get("fanyi", {}).get("tran", "").strip()
    words_count = len(clean_text.split())
    final_defs = []
    pos_str = "-"

    if words_count == 1 and "-" not in clean_text:
        pos_list = []
        for line in raw_trans:
            m = re.match(r"^([a-zA-Z]+\.)\s*(.+)$", line)
            if m:
                pos = m.group(1)
                if pos not in pos_list: pos_list.append(pos)
                final_defs.append(f"- **{pos}** {m.group(2)}")
            else:
                final_defs.append(f"- {line}")
        if not final_defs and web_trans_list:
            for w in web_trans_list[:3]:
                final_defs.append(f"- {w}")
        if not final_defs and fanyi_text:
            final_defs.append(f"- {fanyi_text}")
        pos_str = " / ".join(pos_list) if pos_list else "-"
    else:
        pos_str = "[复合词]" if "-" in clean_text else "[短语/句子]"
        if raw_trans: final_defs.extend([f"- {x}" for x in raw_trans])
        if web_trans_list:
            for w in web_trans_list:
                if w not in final_defs: final_defs.append(f"- {w}")
        if not final_defs and fanyi_text:
            final_defs.append(f"- {fanyi_text}")

    is_uncertain = (phonetic_str == "-" and pos_str == "-" and words_count == 1)
    base_form, inflection_type = extract_base_form(data, raw_trans, clean_text)
    verb_forms_str = extract_verb_forms(data)
    memory_str = extract_associative_memory(data, clean_text)

    return {
        "word": raw_text,
        "phonetic": phonetic_str,
        "pos": pos_str,
        "definition": "\n".join(final_defs) if final_defs else "-",
        "is_uncertain": is_uncertain,
        "base_form": base_form,
        "inflection_type": inflection_type,
        "verb_forms": verb_forms_str,
        "memory": memory_str
    }


# --- UI Application ---

class FloatingLookupApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("PaperLex(人工智能论文方向)")
        self.root.geometry("440x420")
        self.root.minsize(340, 260)
        self.root.attributes("-topmost", True)
        self.root.configure(bg="#F8FAFC")

        self.current_info = None
        self.timer_running = False
        self.timer_seconds = 0
        self.deepseek_req_id = 0

        self._build_ui()
        self._bind_events()

    def _build_ui(self):
        top_frame = tk.Frame(self.root, bg="#F8FAFC", pady=4, padx=8)
        top_frame.pack(fill="x", side="top")

        self.entry_var = tk.StringVar()
        self.search_entry = tk.Entry(
            top_frame,
            textvariable=self.entry_var,
            font=("微软雅黑", 9),
            relief="flat",
            highlightthickness=1,
            highlightbackground="#CBD5E1",
            highlightcolor="#2563EB",
            bg="#FFFFFF"
        )
        self.search_entry.pack(side="left", fill="x", expand=True, ipady=2, padx=(0, 6))

        btn_search = tk.Button(
            top_frame,
            text="查词",
            font=("微软雅黑", 8, "bold"),
            bg="#2563EB",
            fg="white",
            relief="flat",
            activebackground="#1D4ED8",
            activeforeground="white",
            cursor="hand2",
            padx=8,
            pady=1,
            command=self.on_search_click
        )
        btn_search.pack(side="right")

        bottom_frame = tk.Frame(self.root, bg="#F1F5F9", pady=5, padx=8)
        bottom_frame.pack(fill="x", side="bottom")

        self.lbl_status = tk.Label(bottom_frame, text="[Enter] 存 | [Esc] 隐", font=("微软雅黑", 8), fg="#64748B",
                                   bg="#F1F5F9")
        self.lbl_status.pack(side="left")

        btn_group_frame = tk.Frame(bottom_frame, bg="#F1F5F9")
        btn_group_frame.pack(side="right")

        btn_cfg = tk.Button(
            btn_group_frame,
            text="设置",
            font=("微软雅黑", 8),
            bg="#E2E8F0",
            fg="#475569",
            relief="flat",
            cursor="hand2",
            padx=6,
            pady=2,
            command=lambda: open_settings_dialog(self.root)
        )
        btn_cfg.pack(side="left", padx=(0, 4))

        self.btn_deepseek = tk.Button(
            btn_group_frame,
            text="深度解析",
            font=("微软雅黑", 8, "bold"),
            bg="#6366F1",
            fg="white",
            relief="flat",
            activebackground="#4F46E5",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=2,
            command=self.on_deepseek_click
        )
        self.btn_deepseek.pack(side="left", padx=(0, 4))

        self.btn_save = tk.Button(
            btn_group_frame,
            text="保存 (Enter)",
            font=("微软雅黑", 8, "bold"),
            bg="#10B981",
            fg="white",
            relief="flat",
            activebackground="#059669",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=2,
            command=self.save_current_word
        )
        self.btn_save.pack(side="left")

        self.header_frame = tk.Frame(self.root, bg="#F8FAFC", padx=8)
        self.header_frame.pack(fill="x", side="top", pady=(1, 2))

        self.lbl_word = tk.Label(self.header_frame, text="就绪", font=("微软雅黑", 11, "bold"), fg="#1E293B",
                                 bg="#F8FAFC")
        self.lbl_word.pack(anchor="w")

        self.lbl_sub = tk.Label(self.header_frame, text="划选按 F8 查词，或在上方打字", font=("微软雅黑", 8),
                                fg="#64748B", bg="#F8FAFC", justify="left")
        self.lbl_sub.pack(anchor="w")

        self.lemma_frame = tk.Frame(self.root, bg="#EFF6FF", padx=8, pady=2)
        self.lbl_lemma_hint = tk.Label(self.lemma_frame, text="", font=("微软雅黑", 8), fg="#1D4ED8", bg="#EFF6FF")
        self.lbl_lemma_hint.pack(side="left")
        self.btn_lemma_jump = tk.Button(
            self.lemma_frame,
            text="",
            font=("微软雅黑", 8, "bold"),
            bg="#2563EB",
            fg="white",
            relief="flat",
            activebackground="#1D4ED8",
            activeforeground="white",
            cursor="hand2",
            padx=6,
            pady=0
        )
        self.btn_lemma_jump.pack(side="right")

        content_frame = tk.Frame(self.root, bg="#F8FAFC", padx=8)
        content_frame.pack(fill="both", expand=True, side="top", pady=2)

        self.cards_frame = tk.Frame(content_frame, bg="#F8FAFC")
        self.txt_frame = tk.Frame(content_frame, bg="#FFFFFF")
        self.txt_frame.pack(fill="both", expand=True)

        self.scrollbar = tk.Scrollbar(self.txt_frame)
        self.scrollbar.pack(side="right", fill="y")

        self.txt_def = tk.Text(
            self.txt_frame,
            font=("微软雅黑", 8),
            bg="#FFFFFF",
            fg="#334155",
            relief="flat",
            highlightthickness=1,
            highlightbackground="#E2E8F0",
            wrap="word",
            padx=8,
            pady=6,
            yscrollcommand=self.scrollbar.set
        )
        self.txt_def.pack(side="left", fill="both", expand=True)
        self.scrollbar.config(command=self.txt_def.yview)

    def _bind_events(self):
        self.search_entry.bind("<Return>", lambda e: self.on_search_click())
        self.root.bind("<Return>", lambda e: self.save_current_word())
        self.root.bind("<Escape>", lambda e: self.hide_window())

    def show_and_focus(self, focus_entry=False):
        self.root.deiconify()
        self.root.attributes("-topmost", True)
        self.root.lift()
        if focus_entry:
            self.search_entry.focus_set()
            self.search_entry.select_range(0, tk.END)

    def hide_window(self):
        self.root.withdraw()

    def set_loading(self, target_text, custom_msg="正在查询..."):
        self.show_and_focus(focus_entry=False)
        self.entry_var.set(target_text)
        self.lbl_word.config(text=target_text, fg="#1E293B")
        self.lbl_sub.config(text=custom_msg, fg="#2563EB")
        self.lemma_frame.pack_forget()
        self.txt_frame.pack(fill="both", expand=True)
        self.cards_frame.pack_forget()
        self.txt_def.delete("1.0", tk.END)
        self.lbl_status.config(text="正在查询...", fg="#2563EB")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def display_english_result(self, info, suggestions=None):
        self.timer_running = False
        self.current_info = info
        self.lbl_word.config(text=info["word"], fg="#1E293B")

        sub_parts = []
        if info.get("pos") and info["pos"] != "-":
            sub_parts.append(info["pos"])
        if info.get("phonetic") and info["phonetic"] != "-":
            sub_parts.append(info["phonetic"])
        sub_line = "   ".join(sub_parts) if sub_parts else "-"

        if info.get("verb_forms"):
            display_sub = f"{sub_line}\n{info['verb_forms']}"
        else:
            display_sub = sub_line

        self.lbl_sub.config(text=display_sub, fg="#0D9488")

        if info.get("base_form"):
            base_w = info["base_form"]
            inf_type = info.get("inflection_type") or "变形"
            self.lbl_lemma_hint.config(text=f"属「{inf_type}」，原词:")
            self.btn_lemma_jump.config(
                text=f"查原词 {base_w}",
                command=lambda target=base_w: self.start_lookup(target)
            )
            self.lemma_frame.pack(fill="x", side="top", padx=8, pady=(0, 2), before=self.txt_frame.master)
        else:
            self.lemma_frame.pack_forget()

        self.cards_frame.pack_forget()
        self.txt_frame.pack(fill="both", expand=True)
        self.txt_def.delete("1.0", tk.END)

        display_text = info["definition"]

        if info.get("memory"):
            display_text += "\n\n---\n" + info["memory"]

        if suggestions:
            display_text += "\n\n---\n### 相近推荐\n"
            for idx, s in enumerate(suggestions, 1):
                exp = f" ({s['explain']})" if s['explain'] else ""
                display_text += f"- **{s['word']}**{exp}\n"

        # 核心：使用修复后的解析器渲染 Markdown、原生表格与高清 LaTeX
        render_markdown_with_latex(self.txt_def, display_text)

        if is_word_already_saved(info["word"]):
            self.lbl_status.config(text="已在生词本中", fg="#D97706")
        else:
            self.lbl_status.config(text="按 [Enter] 保存", fg="#64748B")
        self.btn_save.config(state="normal", text="保存 (Enter)", bg="#10B981")

    def display_disambiguation_options(self, target_text: str, candidates: list):
        """展示多义缩写方向卡片"""
        self.timer_running = False
        self.lemma_frame.pack_forget()
        self.lbl_word.config(text=f"{target_text} (多义学术缩写)", fg="#1E293B")
        self.lbl_sub.config(text="💡 检测到多个主流学术方向，请点击选择精准解析：", fg="#2563EB")

        self.txt_frame.pack_forget()
        self.cards_frame.pack(fill="both", expand=True)
        for widget in self.cards_frame.winfo_children():
            widget.destroy()

        for idx, item in enumerate(candidates, 1):
            full_name = item.get("full_name", "").strip()
            domain = item.get("domain", "").strip()
            btn_text = f"[{idx}] {full_name}\n     领域: {domain}"
            btn = tk.Button(
                self.cards_frame,
                text=btn_text,
                font=("微软雅黑", 8, "bold"),
                bg="#FFFFFF",
                fg="#1E293B",
                relief="flat",
                anchor="w",
                justify="left",
                padx=8,
                pady=4,
                cursor="hand2",
                command=lambda fn=full_name, dm=domain: self.start_deepseek_direct_lookup(
                    target_text,
                    context_hint=f"{target_text} - {fn} ({dm})",
                    skip_disambiguation=True
                )
            )
            btn.pack(fill="x", pady=2)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#EFF6FF", fg="#2563EB"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#FFFFFF", fg="#1E293B"))

        self.lbl_status.config(text="点击具体方向开启定向深度推演", fg="#2563EB")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def display_spelling_suggestions(self, wrong_text, suggestions):
        self.timer_running = False
        self.lemma_frame.pack_forget()
        self.lbl_word.config(text=wrong_text, fg="#DC2626")
        self.lbl_sub.config(text="疑似拼写错误，点击切换：", fg="#D97706")

        self.txt_frame.pack_forget()
        self.cards_frame.pack(fill="both", expand=True)
        for widget in self.cards_frame.winfo_children():
            widget.destroy()

        for idx, item in enumerate(suggestions, 1):
            w = item["word"]
            exp = item["explain"] or "无释义"
            btn = tk.Button(
                self.cards_frame,
                text=f"[{idx}] {w} ({exp})",
                font=("微软雅黑", 8, "bold"),
                bg="#FFFFFF",
                fg="#1E293B",
                relief="flat",
                anchor="w",
                padx=6,
                pady=3,
                cursor="hand2",
                command=lambda target=w: self.start_lookup(target)
            )
            btn.pack(fill="x", pady=2)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#FEF3C7", fg="#D97706"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#FFFFFF", fg="#1E293B"))

        self.lbl_status.config(text="点击推荐词或点右下 深度解析", fg="#D97706")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def display_chinese_candidates(self, chinese_text, candidates):
        self.timer_running = False
        self.lemma_frame.pack_forget()
        self.lbl_word.config(text=chinese_text, fg="#1E293B")
        self.lbl_sub.config(text="推荐英文表达 (点击选择):", fg="#2563EB")

        self.txt_frame.pack_forget()
        self.cards_frame.pack(fill="both", expand=True)
        for widget in self.cards_frame.winfo_children():
            widget.destroy()

        for idx, item in enumerate(candidates, 1):
            eng = item["english"]
            exp = item["explain"]
            btn = tk.Button(
                self.cards_frame,
                text=f"[{idx}] {eng} ({exp})",
                font=("微软雅黑", 8),
                bg="#FFFFFF",
                fg="#1E293B",
                relief="flat",
                anchor="w",
                padx=6,
                pady=3,
                cursor="hand2",
                command=lambda w=eng: self.start_lookup(w)
            )
            btn.pack(fill="x", pady=1)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#EFF6FF", fg="#2563EB"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#FFFFFF", fg="#1E293B"))

        self.lbl_status.config(text="点击选项获取完整释义", fg="#64748B")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def on_search_click(self):
        text = self.entry_var.get().strip()
        if text:
            self.start_lookup(text)

    def on_deepseek_click(self):
        cfg = load_or_init_config()
        api_key = cfg.get("api_key", "").strip()

        if not api_key or "sk-" not in api_key or "your" in api_key or "你的" in api_key:
            open_settings_dialog(self.root)
            cfg = load_or_init_config()
            if not cfg.get("api_key", "").startswith("sk-"):
                return

        target = self.entry_var.get().strip()
        if not target and self.current_info:
            target = self.current_info["word"]
        if target:
            self.start_deepseek_direct_lookup(target)

    def start_lookup(self, text: str):
        self.timer_running = False
        self.set_loading(text)
        threading.Thread(target=self._async_lookup_worker, args=(text,), daemon=True).start()

    def start_deepseek_direct_lookup(self, text: str, context_hint: str = "", skip_disambiguation: bool = False):
        self.deepseek_req_id += 1
        req_id = self.deepseek_req_id

        self.timer_running = True
        self.timer_seconds = 0

        msg = "正在进行学术消歧嗅探..." if not skip_disambiguation else f"正在深度推演「{context_hint or text}」..."
        self.set_loading(text, custom_msg=f"{msg} (0s / 60s)")
        self._tick_deepseek_timer(req_id, custom_label=msg)

        threading.Thread(
            target=self._async_deepseek_worker,
            args=(text, req_id, context_hint, skip_disambiguation),
            daemon=True
        ).start()

    def _tick_deepseek_timer(self, req_id: int, custom_label: str = "正在解析..."):
        if not self.timer_running or req_id != self.deepseek_req_id:
            return

        self.timer_seconds += 1

        if self.timer_seconds > 60:
            self.timer_running = False
            self.lbl_sub.config(text="请求超时 (60s)，已停止", fg="#DC2626")
            self.lbl_status.config(text="超时停止", fg="#DC2626")
            self.txt_def.delete("1.0", tk.END)
            self.txt_def.insert(
                "1.0",
                "请求超过 60 秒未响应，已自动中止。\n\n"
                "可能原因：\n"
                "1. 网络连接不稳定；\n"
                "2. 接口负载过高；\n"
                "3. API Key 或配置无效。"
            )
            self.btn_save.config(state="disabled", bg="#94A3B8")
            return

        self.lbl_sub.config(text=f"{custom_label} ({self.timer_seconds}s / 60s)", fg="#2563EB")
        self.root.after(1000, lambda: self._tick_deepseek_timer(req_id, custom_label))

    def _async_lookup_worker(self, text: str):
        if is_contains_chinese(text):
            candidates = fetch_english_candidates_from_chinese(text)
            self.root.after(0, lambda: self.display_chinese_candidates(text, candidates))
        else:
            info = fetch_entry_info(text)
            suggestions = []
            if not info or info.get("is_uncertain"):
                suggestions = fetch_suggestions(text)

            if not info:
                if suggestions:
                    self.root.after(0, lambda: self.display_spelling_suggestions(text, suggestions))
                else:
                    self.root.after(0, lambda: self._show_not_found(text))
            else:
                self.root.after(0, lambda: self.display_english_result(info, suggestions))

    def _async_deepseek_worker(self, text: str, req_id: int, context_hint: str, skip_disambiguation: bool):
        if not skip_disambiguation:
            candidates = probe_llm_disambiguation(text)
            if req_id != self.deepseek_req_id or not self.timer_running:
                return
            if len(candidates) > 1:
                self.root.after(0, lambda: self.display_disambiguation_options(text, candidates))
                return

        base_info = fetch_entry_info(text)
        phonetic = base_info.get("phonetic", "-") if base_info else "-"
        verb_forms = base_info.get("verb_forms", "") if base_info else ""
        ai_analysis = call_llm_direct_query(text, context_hint=context_hint)

        if req_id != self.deepseek_req_id or not self.timer_running:
            return

        self.timer_running = False

        display_title = context_hint if context_hint else text
        info = {
            "word": display_title,
            "phonetic": phonetic,
            "pos": "[学术解析]",
            "definition": ai_analysis,
            "is_uncertain": False,
            "base_form": None,
            "inflection_type": None,
            "verb_forms": verb_forms,
            "memory": ""
        }
        self.root.after(0, lambda: self.display_english_result(info))

    def _show_not_found(self, text: str):
        self.lbl_word.config(text=text, fg="#DC2626")
        self.lbl_sub.config(text="词典未收录，可尝试 深度解析", fg="#DC2626")
        self.lbl_status.config(text="未收录", fg="#DC2626")
        self.btn_save.config(state="disabled", bg="#94A3B8")

    def save_current_word(self):
        if not self.current_info:
            return
        success = save_to_excel(self.current_info)
        if success:
            self.lbl_status.config(text="已保存至表格", fg="#059669")
            self.btn_save.config(text="已保存", state="disabled", bg="#94A3B8")
        else:
            self.lbl_status.config(text="保存失败：请检查 Excel 是否被占用", fg="#DC2626")

    def run(self):
        self.root.mainloop()


# --- Global Listener ---

app_instance = None
is_f8_busy = False


def trigger_f8_hotkey():
    global is_f8_busy, app_instance
    if is_f8_busy or not app_instance:
        return
    is_f8_busy = True
    try:
        pyperclip.copy("")
        time.sleep(0.03)

        keyboard.send("ctrl+c")

        selected_text = ""
        for _ in range(8):
            time.sleep(0.04)
            t = pyperclip.paste().strip()
            if t:
                selected_text = t
                break

        if selected_text:
            app_instance.root.after(0, lambda: app_instance.start_lookup(selected_text))
        else:
            app_instance.root.after(0, lambda: app_instance.show_and_focus(focus_entry=True))
    finally:
        is_f8_busy = False


def start_keyboard_listener():
    keyboard.add_hotkey("f8", trigger_f8_hotkey)
    keyboard.wait()


def main():
    global app_instance
    init_excel()
    load_or_init_config()

    print(f"Loaded config: {CONFIG_FILE}")
    print("Service running. Hotkey [F8] enabled.")

    t = threading.Thread(target=start_keyboard_listener, daemon=True)
    t.start()

    app_instance = FloatingLookupApp()
    app_instance.run()


if __name__ == "__main__":
    main()