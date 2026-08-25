import os
import re
import time
import datetime
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

TARGET_FILE = "words.xlsx"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}
# 禁用 requests 的环境变量代理，直连有道服务器，杜绝卡顿挂起
NO_PROXY = {
    "http": None,
    "https": None
}


def init_target_excel():
    """初始化目标表格"""
    headers = ["单词/内容", "音标", "词性/类型", "完整释义/译文", "记录时间"]
    if not os.path.exists(TARGET_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "生词本"
        ws.append(headers)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 65
        ws.column_dimensions["E"].width = 18
        wb.save(TARGET_FILE)


def get_existing_items():
    """读取已保存的词条，实现断点续传"""
    if not os.path.exists(TARGET_FILE):
        return set()
    try:
        wb = load_workbook(TARGET_FILE, read_only=True)
        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row and row[0]:
                existing.add(str(row[0]).strip().lower())
        wb.close()
        return existing
    except Exception:
        return set()


def extract_nested_trans(node):
    results = []
    if isinstance(node, dict):
        if "pos" in node and "tran" in node:
            results.append(f"{node['pos']} {node['tran']}")
        elif "i" in node:
            if isinstance(node["i"], list):
                for item in node["i"]:
                    if isinstance(item, str) and item.strip():
                        results.append(item.strip())
                    elif isinstance(item, dict):
                        results.extend(extract_nested_trans(item))
            elif isinstance(node["i"], str) and node["i"].strip():
                results.append(node["i"].strip())
        elif "tran" in node and isinstance(node["tran"], str):
            results.append(node["tran"].strip())
        else:
            for v in node.values():
                results.extend(extract_nested_trans(v))
    elif isinstance(node, list):
        for item in node:
            results.extend(extract_nested_trans(item))
    return results


def query_youdao(raw_text: str):
    """底层查询请求"""
    url = "https://dict.youdao.com/jsonapi"
    try:
        response = requests.get(
            url,
            params={"client": "mobile", "q": raw_text},
            headers=HEADERS,
            proxies=NO_PROXY,  # 强制直连
            timeout=4
        )
        if response.status_code == 200:
            return response.json()
    except Exception:
        pass
    return None


def fetch_entry_info(text: str):
    """智能清洗与多通道识别"""
    clean_text = text.strip()
    if not clean_text:
        return None

    # 尝试原词和去除括号后的小写词
    cleaned_candidates = [clean_text]
    # 如果含有括号如 Kinda (kind of)，拆出主词和括号内词
    if "(" in clean_text or "（" in clean_text:
        no_bracket = re.sub(r'[\(（].*?[\)）]', '', clean_text).strip()
        if no_bracket and no_bracket not in cleaned_candidates:
            cleaned_candidates.append(no_bracket)
    if clean_text.lower() not in cleaned_candidates:
        cleaned_candidates.append(clean_text.lower())

    data = None
    for cand in cleaned_candidates:
        data = query_youdao(cand)
        if data and (data.get("ec") or data.get("fanyi") or data.get("web_trans") or data.get("simple")):
            break

    if not data:
        return None

    ec_data = data.get("ec", {}).get("word", [{}])[0] if data.get("ec") else {}
    simple_data = data.get("simple", {}).get("word", [{}])[0] if data.get("simple") else {}

    # 1. 音标
    uk_phone = ec_data.get("ukphone") or simple_data.get("ukphone", "")
    us_phone = ec_data.get("usphone") or simple_data.get("usphone", "")
    phonetic_parts = []
    if uk_phone: phonetic_parts.append(f"英 /{uk_phone}/")
    if us_phone: phonetic_parts.append(f"美 /{us_phone}/")
    if not phonetic_parts:
        generic_phone = ec_data.get("phone") or simple_data.get("phone", "")
        if generic_phone: phonetic_parts.append(f"/{generic_phone}/")
    phonetic_str = "  ".join(phonetic_parts) if phonetic_parts else "-"

    # 2. 词典释义
    raw_trans = []
    for tr in ec_data.get("trs", []):
        for item in extract_nested_trans(tr):
            if item and item not in raw_trans:
                raw_trans.append(item)

    # 3. 网络短语释义
    web_trans_list = []
    for item in data.get("web_trans", {}).get("web-translation", []):
        for t in item.get("trans", []):
            val = t.get("value", "").strip()
            if val and val not in web_trans_list and val not in raw_trans:
                web_trans_list.append(val)

    # 4. 翻译引擎
    fanyi_text = data.get("fanyi", {}).get("tran", "").strip()

    # 5. 分流组装
    words_count = len(clean_text.split())
    final_defs = []

    if words_count == 1:
        pos_list = []
        for line in raw_trans:
            m = re.match(r"^([a-zA-Z]+\.)\s*(.+)$", line)
            if m:
                pos = m.group(1)
                if pos not in pos_list: pos_list.append(pos)
                final_defs.append(f"{pos} {m.group(2)}")
            else:
                final_defs.append(line)
        if not final_defs and web_trans_list:
            final_defs.extend(web_trans_list[:3])
        if not final_defs and fanyi_text:
            final_defs.append(fanyi_text)
        pos_str = " / ".join(pos_list) if pos_list else "-"
    elif words_count <= 5 and not re.search(r'[.!?。！？]$', clean_text):
        pos_str = "[短语]"
        if raw_trans:
            final_defs.extend(raw_trans)
        if web_trans_list:
            for w in web_trans_list:
                if w not in final_defs:
                    final_defs.append(w)
        if not final_defs and fanyi_text:
            final_defs.append(fanyi_text)
    else:
        pos_str = "[句子]"
        phonetic_str = "-"
        if fanyi_text:
            final_defs.append(fanyi_text)
        elif raw_trans:
            final_defs.extend(raw_trans)
        elif web_trans_list:
            final_defs.append("；".join(web_trans_list[:3]))

    if not final_defs and phonetic_str == "-":
        return None

    return {
        "word": clean_text,
        "phonetic": phonetic_str,
        "pos": pos_str,
        "definition": "\n".join(final_defs) if final_defs else "-"
    }


def start_import(old_file_path: str):
    if not os.path.exists(old_file_path):
        print(f"❌ 找不到文件: {old_file_path}")
        return

    init_target_excel()
    existing_items = get_existing_items()

    print(f"📖 正在扫描旧表格「{old_file_path}」第一列...")
    try:
        wb_old = load_workbook(old_file_path, data_only=True)
        ws_old = wb_old.active
    except Exception as e:
        print(f"❌ 无法打开旧表格: {e}")
        return

    items_to_query = []
    # 自动忽略的表头模式
    header_pattern = r"(词条|单词|生词|word|term|vocabulary|phrase|sentence)"

    for row in ws_old.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        val = str(row[0]).strip()
        if not val or len(val) <= 1:
            continue

        # 过滤包含“词条/单词/Term”等表头字样的行
        if re.search(header_pattern, val, re.IGNORECASE) and len(val.split()) <= 3 and (
                "(" in val or "/" in val or len(val) <= 10):
            continue

        if val.lower() in existing_items:
            continue
        if val not in items_to_query:
            items_to_query.append(val)

    total = len(items_to_query)
    if total == 0:
        print("💡 所有内容均已存在于 words.xlsx 中（无新内容需导入）。")
        return

    print(f"✨ 共发现 {total} 条未导入内容，开始批量处理（按 Ctrl+C 可随时安全退出）...\n")

    success_count = 0
    fail_count = 0

    try:
        for idx, item in enumerate(items_to_query, start=1):
            display_item = item if len(item) <= 25 else item[:22] + "..."
            print(f"[{idx}/{total}] 查询: {display_item:<26} ... ", end="", flush=True)

            info = fetch_entry_info(item)
            if info:
                now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
                try:
                    wb_tgt = load_workbook(TARGET_FILE)
                    ws_tgt = wb_tgt.active
                    ws_tgt.append([
                        info["word"],
                        info["phonetic"],
                        info["pos"],
                        info["definition"],
                        now
                    ])
                    r = ws_tgt.max_row
                    ws_tgt.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="center")
                    ws_tgt.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="center")
                    for c in [2, 3, 5]:
                        ws_tgt.cell(row=r, column=c).alignment = Alignment(vertical="center")
                    wb_tgt.save(TARGET_FILE)
                    success_count += 1
                    print("✅ 成功")
                except PermissionError:
                    print("⚠️ 保存被拦截（请关闭 words.xlsx）")
            else:
                fail_count += 1
                print("❌ 未查到")

            time.sleep(0.2)
    except KeyboardInterrupt:
        print("\n\n⚠️ 已检测到手动中断！已查询的数据已安全保存在 words.xlsx 中。")

    print("\n" + "=" * 50)
    print(f"🎉 处理结束！本次新增导入: {success_count} 条，未查到: {fail_count} 条")
    print(f"📁 目标文件: {TARGET_FILE}")
    print("=" * 50)


if __name__ == "__main__":
    old_file = input("请输入旧 Excel 表格的文件名（例如 words0.xlsx）: ").strip()
    if old_file:
        start_import(old_file)